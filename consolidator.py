"""Consolidator — turn raw account inputs into per-year, per-month workbooks.

For each account configured in `config/accounts.yaml`:
  1. Scan `Transactions/{Account}/` for statement source files.
  2. Scan `Reports/Karibu/{Account}/` for Karibu ledger CSVs.
  3. Run the configured parser on every file (always — the state cache at
     `~/.local/share/BSR_Recon/state/consolidator_state.json` is for
     performance, not correctness; a full re-scan is mandatory per spec
     §6.7 so the user can never get stuck on stale data).
  4. Dedupe (statements: `(date, txn_id, amount, direction)`; Karibu:
     `(date, narration, dr, cr, balance)`).
  5. Split by year, then by month.
  6. Write `{Account} Transactions - {YYYY}.xlsx` and
     `{Account} Karibu Ledger - {YYYY}.xlsx` to `Statements/{Account}/`.
     Sheets: `Jan`..`Dec` for months in the data range. Unparseable-date
     rows go to a separate `Unparseable` sheet (Joash's decision —
     surface for review rather than drop silently).

The consolidator **rebuilds from source CSVs only**. It does not merge
against a previous output workbook. This is the structural fix for the
April-6 NaT-baseline bug (see BUGFIX.md).
"""

from __future__ import annotations

import calendar
import hashlib
import json
import logging
import zipfile
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import AccountConfig
from parsers import NormalizedRecord, get_parser
from utils.safe_write import check_xlsx_lock


logger = logging.getLogger("bsr_recon.consolidator")

# BSR brand colours.
BSR_DARK_GREEN = "1A4D2E"   # primary
BSR_GOLD = "B8922A"         # amount column
BSR_VERY_LIGHT_GREEN = "EAF3EC"  # alternating data row fill
BSR_WHITE = "FFFFFF"

MONTH_ABBREV = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

STATE_FILENAME = "consolidator_state.json"


@dataclass
class ConsolidateResult:
    account: str
    statement_files_seen: int
    karibu_files_seen: int
    statement_records_total: int
    statement_records_unique: int
    karibu_records_total: int
    karibu_records_unique: int
    statement_workbooks_written: list[Path]
    karibu_workbooks_written: list[Path]
    statement_unparseable: int
    karibu_unparseable: int

    @property
    def latest_statement_workbook(self) -> Path | None:
        # Unparseable rows are parked in the latest-year workbook (see
        # _write_yearly_workbooks). This pointer lets the caller name the
        # exact file in the run summary.
        return self.statement_workbooks_written[-1] if self.statement_workbooks_written else None

    @property
    def latest_karibu_workbook(self) -> Path | None:
        return self.karibu_workbooks_written[-1] if self.karibu_workbooks_written else None


# ---------------------------------------------------------------------------
# Public entrypoint
# ---------------------------------------------------------------------------

def consolidate_account(account: AccountConfig, base_dir: Path) -> ConsolidateResult:
    """Run the consolidator for a single account. See module docstring."""
    base = Path(base_dir)
    tx_dir = base / "Transactions" / account.name
    karibu_dir = base / "Reports" / "Karibu" / account.name
    out_dir = base / "Statements" / account.name
    out_dir.mkdir(parents=True, exist_ok=True)

    # ---- Statements ----
    tx_records = _parse_directory(
        tx_dir, account.statement_parser,
        state_dir=base / "state",
    )
    tx_unique = _dedupe(tx_records, _statement_dedup_key)
    tx_unparseable = sum(1 for r in tx_unique if r.date is None)
    statement_paths = _write_yearly_workbooks(
        records=tx_unique,
        out_dir=out_dir,
        filename_prefix=f"{account.name} Transactions",
        kind="statement",
    )

    # ---- Karibu ----
    karibu_records = _parse_directory(
        karibu_dir, account.karibu_parser,
        state_dir=base / "state",
        parser_kwargs={"karibu_account": account.karibu_account} if account.karibu_account else None,
    )
    karibu_unique = _dedupe(karibu_records, _karibu_dedup_key)
    karibu_unparseable = sum(1 for r in karibu_unique if r.date is None)
    karibu_paths = _write_yearly_workbooks(
        records=karibu_unique,
        out_dir=out_dir,
        filename_prefix=f"{account.name} Karibu Ledger",
        kind="karibu",
    )

    return ConsolidateResult(
        account=account.name,
        statement_files_seen=_count_files(tx_dir),
        karibu_files_seen=_count_files(karibu_dir),
        statement_records_total=len(tx_records),
        statement_records_unique=len(tx_unique),
        karibu_records_total=len(karibu_records),
        karibu_records_unique=len(karibu_unique),
        statement_workbooks_written=statement_paths,
        karibu_workbooks_written=karibu_paths,
        statement_unparseable=tx_unparseable,
        karibu_unparseable=karibu_unparseable,
    )


# ---------------------------------------------------------------------------
# Parsing + state cache (performance only — full re-scan is mandatory)
# ---------------------------------------------------------------------------

def _parse_directory(
    directory: Path,
    parser_name: str,
    state_dir: Path,
    parser_kwargs: dict | None = None,
) -> list[NormalizedRecord]:
    """Parse every supported file in `directory` and return all records.

    Always parses every file (the spec is explicit: re-scan every run).
    The state cache is updated to record the fingerprint of each file so a
    future variant can skip unchanged files — but Phase 2 always re-scans.
    """
    if not directory.is_dir():
        return []
    parse_fn = get_parser(parser_name)
    parser_kwargs = parser_kwargs or {}

    state = _load_state(state_dir)
    files = sorted(_supported_files(directory))
    records: list[NormalizedRecord] = []

    for f in files:
        try:
            chunk = parse_fn(f, **parser_kwargs)
        except Exception as exc:
            logger.error("parser %s failed on %s: %s", parser_name, f.name, exc)
            continue
        records.extend(chunk)
        state[str(f)] = {
            "fingerprint": _fingerprint(f),
            "parsed_records": len(chunk),
            "parsed_at": datetime.now().isoformat(timespec="seconds"),
        }

    _save_state(state_dir, state)
    return records


def _supported_files(directory: Path) -> Iterable[Path]:
    for p in directory.iterdir():
        if not p.is_file():
            continue
        if p.name.startswith("."):
            continue
        if p.suffix.lower() in {".csv", ".xlsx", ".xls"}:
            yield p


def _count_files(directory: Path) -> int:
    if not directory.is_dir():
        return 0
    return sum(1 for _ in _supported_files(directory))


def _fingerprint(path: Path) -> str:
    st = path.stat()
    h = hashlib.sha256()
    h.update(str(path).encode("utf-8"))
    h.update(str(st.st_size).encode("ascii"))
    h.update(str(int(st.st_mtime)).encode("ascii"))
    return h.hexdigest()


def _load_state(state_dir: Path) -> dict:
    f = state_dir / STATE_FILENAME
    if not f.exists():
        return {}
    try:
        return json.loads(f.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return {}


def _save_state(state_dir: Path, state: dict) -> None:
    state_dir.mkdir(parents=True, exist_ok=True)
    f = state_dir / STATE_FILENAME
    f.write_text(json.dumps(state, indent=2, sort_keys=True), encoding="utf-8")


# ---------------------------------------------------------------------------
# Dedup
# ---------------------------------------------------------------------------

def _statement_dedup_key(r: NormalizedRecord) -> tuple:
    # date may be None for unparseable rows; use a sentinel string so they
    # don't collide with valid 0001-01-01 epochs (Python's datetime min).
    d = r.date.isoformat() if r.date else f"_NAT_{r.source_file}_{id(r)}"
    return (d, r.txn_id, str(r.amount), r.direction)


def _karibu_dedup_key(r: NormalizedRecord) -> tuple:
    d = r.date.isoformat() if r.date else f"_NAT_{r.source_file}_{id(r)}"
    return (d, r.counterparty, str(r.amount), r.direction,
            r.raw.get("Balance", ""))


def _dedupe(records: list[NormalizedRecord], key_fn) -> list[NormalizedRecord]:
    """Stable-dedupe — first occurrence wins. Records with date=None are
    intentionally never collapsed against each other (they carry a unique
    sentinel in the key) so the user sees every unparseable row in the
    review sheet, not just one.
    """
    seen = set()
    out: list[NormalizedRecord] = []
    for r in records:
        k = key_fn(r)
        if k in seen:
            continue
        seen.add(k)
        out.append(r)
    # Sort: valid dates ascending, NaT rows last.
    out.sort(key=lambda r: (r.date is None, r.date or datetime.min))
    return out


# ---------------------------------------------------------------------------
# Workbook writing
# ---------------------------------------------------------------------------

def _write_yearly_workbooks(
    records: list[NormalizedRecord],
    out_dir: Path,
    filename_prefix: str,
    kind: str,  # "statement" or "karibu"
) -> list[Path]:
    """Group records by year, write one workbook per year. Returns paths."""
    if not records:
        return []
    by_year: dict[int | None, list[NormalizedRecord]] = {}
    for r in records:
        year = r.date.year if r.date else None
        by_year.setdefault(year, []).append(r)

    written: list[Path] = []
    # Group year=None records together with the smallest valid year's file
    # for visibility, OR write them to a dedicated "Unknown Year" file if
    # there's no valid year at all. Simplest: keep year=None in its own
    # workbook so the user gets one obvious review file.
    valid_years = sorted(y for y in by_year if y is not None)
    for year in valid_years:
        year_records = by_year[year]
        # Add the unparseable rows to every year so they're not lost. Better:
        # write them to the *latest* year so they appear in the most-recent
        # review surface. For now: include them in the latest year only.
        if year == valid_years[-1] and None in by_year:
            year_records = year_records + by_year[None]
        path = out_dir / f"{filename_prefix} - {year}.xlsx"
        _write_year_workbook(year_records, path, kind=kind, year=year)
        written.append(path)

    if not valid_years and None in by_year:
        # Edge case: every record had an unparseable date. Park them in a
        # year-less review workbook for the user.
        path = out_dir / f"{filename_prefix} - UNKNOWN.xlsx"
        _write_year_workbook(by_year[None], path, kind=kind, year=None)
        written.append(path)
    return written


def _write_year_workbook(
    records: list[NormalizedRecord],
    path: Path,
    kind: str,
    year: int | None,
) -> None:
    """Write one year's records into a workbook with Jan..Dec sheets and an
    optional Unparseable sheet for rows whose date couldn't be parsed.
    """
    check_xlsx_lock(path)

    healthy = [r for r in records if r.date is not None]
    bad = [r for r in records if r.date is None]

    columns = _columns_for(kind)
    wb = Workbook()
    # openpyxl creates a default "Sheet"; we'll replace it month-by-month.
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Decide which months to include. Always include any month that has data;
    # for a year-anchored workbook, fill empty months in the data range so
    # layouts look consistent across years that span the same span.
    months_with_data = sorted({r.date.month for r in healthy})
    if months_with_data:
        # Include every month between min and max so users can flip through
        # without missing tabs in the middle.
        months_to_render = list(range(months_with_data[0], months_with_data[-1] + 1))
    else:
        months_to_render = []

    for month in months_to_render:
        sheet_name = MONTH_ABBREV[month - 1]
        ws = wb.create_sheet(sheet_name)
        month_rows = [r for r in healthy if r.date.month == month]
        _write_sheet(ws, columns, month_rows, kind=kind)

    if bad:
        ws = wb.create_sheet("Unparseable")
        _write_sheet(ws, columns, bad, kind=kind, unparseable=True)

    # Edge case: no data at all (should not happen — caller checks).
    if not wb.sheetnames:
        wb.create_sheet("Empty")

    _pin_workbook_metadata(wb)
    wb.save(path)
    _pin_zip_member_timestamps(path)


# A fixed epoch used for workbook `created`/`modified` properties so that
# re-running the consolidator on the same inputs produces byte-identical
# output (spec requirement). The value is arbitrary; it just has to be
# stable across runs.
_FIXED_TS = datetime(2024, 1, 1, 0, 0, 0)


def _pin_workbook_metadata(wb) -> None:
    """Strip non-deterministic timestamps so two runs over identical inputs
    yield byte-identical output."""
    props = wb.properties
    props.creator = "BSR_Recon"
    props.lastModifiedBy = "BSR_Recon"
    props.created = _FIXED_TS
    props.modified = _FIXED_TS


# openpyxl writes each xlsx as a ZIP. Even if all member contents are byte-
# identical, each member's local-file-header timestamp defaults to "now",
# making the overall file hash differ between runs. We post-process: re-zip
# with all entry timestamps pinned to _FIXED_ZIP_TS.
_FIXED_ZIP_TS = (2024, 1, 1, 0, 0, 0)


def _pin_zip_member_timestamps(path: Path) -> None:
    with zipfile.ZipFile(path, "r") as src:
        members = [(info, src.read(info.filename)) for info in src.infolist()]
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as dst:
        for info, data in members:
            new_info = zipfile.ZipInfo(filename=info.filename, date_time=_FIXED_ZIP_TS)
            new_info.compress_type = info.compress_type
            new_info.external_attr = info.external_attr
            new_info.create_system = info.create_system
            dst.writestr(new_info, data)


def _columns_for(kind: str) -> list[tuple[str, str]]:
    """(column_header, value_kind) tuples per row layout."""
    if kind == "statement":
        return [
            ("Date", "date"),
            ("Transaction ID", "text"),
            ("Direction", "text"),
            ("Counterparty", "text"),
            ("Transaction Type", "text"),
            ("Amount (UGX)", "amount"),
            ("Source File", "text"),
            ("Audit Flag", "text"),
        ]
    if kind == "karibu":
        return [
            ("Date", "date"),
            ("Narration", "text"),
            ("Direction", "text"),
            ("Amount (UGX)", "amount"),
            ("DR", "amount"),
            ("CR", "amount"),
            ("Balance", "text"),
            ("Source File", "text"),
            ("Audit Flag", "text"),
        ]
    raise ValueError(f"unknown kind: {kind}")


def _row_values(r: NormalizedRecord, kind: str) -> list:
    if kind == "statement":
        return [
            r.date,
            r.txn_id,
            r.direction,
            r.counterparty,
            r.txn_type,
            float(r.amount) if r.amount is not None else None,
            r.source_file,
            r.audit_flag,
        ]
    return [  # karibu
        r.date,
        r.counterparty,  # = Narration
        r.direction,
        float(r.amount) if r.amount is not None else None,
        float(r.raw.get("DR") or 0),
        float(r.raw.get("CR") or 0),
        r.raw.get("Balance", ""),
        r.source_file,
        r.audit_flag,
    ]


def _write_sheet(ws, columns: list[tuple[str, str]],
                 rows: list[NormalizedRecord], kind: str,
                 unparseable: bool = False) -> None:
    """Write one month or unparseable sheet with BSR styling."""
    header_fill = PatternFill(start_color=BSR_DARK_GREEN, end_color=BSR_DARK_GREEN, fill_type="solid")
    gold_fill = PatternFill(start_color=BSR_GOLD, end_color=BSR_GOLD, fill_type="solid")
    zebra_fill = PatternFill(start_color=BSR_VERY_LIGHT_GREEN, end_color=BSR_VERY_LIGHT_GREEN, fill_type="solid")
    header_font = Font(name="Arial", size=10, bold=True, color=BSR_WHITE)
    data_font = Font(name="Arial", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Header row.
    for col_idx, (header, kind_) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = center
        if header == "Amount (UGX)":
            cell.fill = gold_fill
            cell.font = Font(name="Arial", size=10, bold=True, color=BSR_WHITE)
        else:
            cell.fill = header_fill

    # Data rows.
    for row_idx, record in enumerate(rows, start=2):
        values = _row_values(record, kind)
        for col_idx, ((_, value_kind), v) in enumerate(zip(columns, values), 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            cell.font = data_font
            if (row_idx - 2) % 2 == 1:
                cell.fill = zebra_fill
            if value_kind == "date" and v is not None:
                cell.number_format = "yyyy-mm-dd hh:mm"
            elif value_kind == "amount":
                cell.number_format = "#,##0"

    # Freeze the header row.
    ws.freeze_panes = "A2"

    # Autofilter on the data range.
    if rows:
        last_col_letter = get_column_letter(len(columns))
        ws.auto_filter.ref = f"A1:{last_col_letter}{len(rows) + 1}"

    # Reasonable column widths.
    widths = {"Date": 18, "Transaction ID": 18, "Direction": 10,
              "Counterparty": 28, "Transaction Type": 14, "Amount (UGX)": 14,
              "Source File": 32, "Audit Flag": 22, "Narration": 36,
              "DR": 12, "CR": 12, "Balance": 14}
    for col_idx, (header, _) in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(header, 16)


# ---------------------------------------------------------------------------
# Multi-account driver
# ---------------------------------------------------------------------------

def consolidate_all(accounts: Iterable[AccountConfig], base_dir: Path) -> list[ConsolidateResult]:
    return [consolidate_account(a, base_dir) for a in accounts]
