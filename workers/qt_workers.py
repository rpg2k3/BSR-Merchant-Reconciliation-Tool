"""QThread workers for non-blocking UI operations."""

from pathlib import Path

from PyQt6.QtCore import QThread, pyqtSignal


class UpdateWorker(QThread):
    """Worker thread for statement update operations."""

    log_signal = pyqtSignal(str, str)  # message, level
    finished_signal = pyqtSignal(dict)  # result dict

    def __init__(self, channel: str, base_dir: Path):
        super().__init__()
        self.channel = channel
        self.base_dir = base_dir

    def run(self):
        try:
            from core.updater import update_mtn_statement, update_airtel_statement

            if self.channel == "MTN":
                result = update_mtn_statement(self.base_dir, self._log)
            else:
                result = update_airtel_statement(self.base_dir, self._log)
            self.finished_signal.emit(result)
        except Exception as e:
            self._log(f"Error: {e}", "error")
            self.finished_signal.emit({"error": str(e)})

    def _log(self, msg, level="info"):
        self.log_signal.emit(msg, level)


class ReconcileWorker(QThread):
    """Worker thread for reconciliation operations."""

    log_signal = pyqtSignal(str, str)
    finished_signal = pyqtSignal(dict)
    ai_narrative_signal = pyqtSignal(str)  # AI analysis text

    def __init__(self, channel: str, base_dir: Path, config: dict):
        super().__init__()
        self.channel = channel
        self.base_dir = base_dir
        self.config = config

    def run(self):
        try:
            from core.reconciler import reconcile
            result = reconcile(self.channel, self.base_dir, self.config, self._log)

            # Optional AI analysis
            api_key = self.config.get("claude_api_key", "")
            if api_key and "error" not in result:
                self._log("Running AI audit analysis...", "info")
                try:
                    from core.anomalies import get_flagged_summary
                    from core.ai_analyst import run_ai_analysis, save_narrative

                    # Re-load the reconciliation output to get flagged rows
                    from openpyxl import load_workbook
                    import pandas as pd

                    recon_path = Path(result.get("recon_path", ""))
                    if recon_path.exists():
                        wb = load_workbook(recon_path, data_only=True)
                        karibu_out = _sheet_to_df(wb["Karibu Report"])
                        stmt_out = _sheet_to_df(wb["Merchant Statement"])
                        wb.close()

                        flagged = get_flagged_summary(karibu_out, stmt_out)
                        if flagged:
                            narrative = run_ai_analysis(flagged, self.channel, api_key)
                            if narrative:
                                txt_path = save_narrative(narrative, recon_path, self.channel)
                                self._log(f"AI narrative saved: {txt_path}")
                                self.ai_narrative_signal.emit(narrative)
                        else:
                            self.ai_narrative_signal.emit("No anomalies flagged — clean reconciliation.")
                except Exception as e:
                    self._log(f"AI analysis error: {e}", "warning")

            self.finished_signal.emit(result)
        except Exception as e:
            self._log(f"Error: {e}", "error")
            self.finished_signal.emit({"error": str(e)})

    def _log(self, msg, level="info"):
        self.log_signal.emit(msg, level)


def _sheet_to_df(ws):
    """Convert an openpyxl worksheet to a pandas DataFrame."""
    import pandas as pd
    headers = [cell.value for cell in ws[1]]
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data.append(list(row))
    return pd.DataFrame(data, columns=headers)
