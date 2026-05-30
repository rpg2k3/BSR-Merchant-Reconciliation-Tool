"""Excel writer for the Phase-3 reconciliation workbook.

Layout (matches `samples/BSR_MTN_Reconciliation.xlsx` apart from the
statement sheet being named "Statement" rather than "Merchant Statement"
— the new name generalises to non-merchant accounts like Petty Cash):

  Sheet 1: "Karibu Report"   per-row Karibu ledger with Status / Match Type /
                             Confidence / Matched Ref / Audit Flag / Comments
  Sheet 2: "Statement"       per-row statement transactions with the same
                             status columns
  Sheet 3: "Dashboard"       a thin summary — totals, match counts, audit
                             flag breakdown, contras snapshot

BSR branding:
  - Dark green  #1A4D2E  primary
  - Gold        #B8922A  amount-column accent
  - Mid green   #2D6A4F  used in row tints / dashboard secondary
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from utils.safe_write import check_xlsx_lock


# ---------------------------------------------------------------------------
# BSR palette
# ---------------------------------------------------------------------------
BSR_DARK_GREEN = "1A4D2E"
BSR_GOLD = "B8922A"
BSR_MID_GREEN = "2D6A4F"
BSR_VERY_LIGHT_GREEN = "EAF3EC"
BSR_WHITE = "FFFFFF"

ROW_MATCHED = "D6EFDD"
ROW_NOT_IN_STATEMENT = "FDECEA"
ROW_NOT_IN_KARIBU = "FFF3CD"
ROW_FLAGGED = "FCE4EC"
ROW_AUDIT_BG = "FF9800"

THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)


# ---------------------------------------------------------------------------
# Public entrypoint
# ---------------------------------------------------------------------------

def write_reconciliation_workbook(
    karibu_out: pd.DataFrame,
    stmt_out: pd.DataFrame,
    dashboard_lines: list[str],
    output_path: Path,
) -> Path:
    """Write the three-sheet reconciliation workbook to `output_path`."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    check_xlsx_lock(output_path)

    wb = Workbook()
    ws_k = wb.active
    ws_k.title = "Karibu Report"
    _write_recon_sheet(ws_k, karibu_out)

    ws_s = wb.create_sheet("Statement")
    _write_recon_sheet(ws_s, stmt_out)

    ws_d = wb.create_sheet("Dashboard")
    _write_dashboard(ws_d, dashboard_lines)

    wb.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------

def _write_recon_sheet(ws, df: pd.DataFrame) -> None:
    """Write one Karibu/Statement-style sheet."""
    if df.empty:
        ws.cell(row=1, column=1, value="No data")
        return

    headers = list(df.columns)
    hfont = Font(name="Arial", size=10, bold=True, color=BSR_WHITE)
    hfill = PatternFill(start_color=BSR_DARK_GREEN, end_color=BSR_DARK_GREEN, fill_type="solid")
    gold_fill = PatternFill(start_color=BSR_GOLD, end_color=BSR_GOLD, fill_type="solid")
    zebra = PatternFill(start_color=BSR_VERY_LIGHT_GREEN, end_color=BSR_VERY_LIGHT_GREEN, fill_type="solid")

    flag_fill = PatternFill(start_color=ROW_AUDIT_BG, end_color=ROW_AUDIT_BG, fill_type="solid")
    flag_font = Font(name="Arial", size=9, bold=True, color="000000")
    data_font = Font(name="Arial", size=9)

    amount_headers = {"DR (UGX)", "CR (UGX)", "Amount (UGX)", "Balance"}
    date_headers = {"Date"}

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = hfont
        cell.fill = gold_fill if h in {"DR (UGX)", "CR (UGX)", "Amount (UGX)"} else hfill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        status = str(row.get("Status", ""))
        flag = str(row.get("Audit Flag", "") or "")
        has_flag = bool(flag) and flag not in {"nan", "—"}
        bg = _row_fill(status, has_flag)

        for col_idx, (col_name, val) in enumerate(zip(headers, row), 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val if not pd.isna(val) else "")
            cell.border = THIN_BORDER
            if bg is not None:
                cell.fill = bg
            elif (row_idx - 2) % 2 == 1:
                cell.fill = zebra

            if col_name == "Status":
                cell.font = _status_font(str(val) if not pd.isna(val) else "")
            elif col_name == "Confidence":
                cell.font = _confidence_font(str(val) if not pd.isna(val) else "")
            elif col_name == "Audit Flag" and has_flag:
                cell.fill = flag_fill
                cell.font = flag_font
            else:
                cell.font = data_font

            if col_name in amount_headers:
                cell.number_format = "#,##0"
            elif col_name in date_headers:
                cell.number_format = "DD/MM/YYYY"

    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"
    _auto_width(ws)


def _write_dashboard(ws, lines: list[str]) -> None:
    title_font = Font(name="Arial", size=14, bold=True, color=BSR_DARK_GREEN)
    section_font = Font(name="Arial", size=11, bold=True, color=BSR_MID_GREEN)
    data_font = Font(name="Arial", size=10)

    for i, line in enumerate(lines, 1):
        cell = ws.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = title_font
        elif line and not line.startswith(" "):
            cell.font = section_font
        else:
            cell.font = data_font

    ws.column_dimensions["A"].width = 65


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

def _row_fill(status: str, has_flag: bool) -> PatternFill | None:
    if has_flag:
        return PatternFill(start_color=ROW_FLAGGED, end_color=ROW_FLAGGED, fill_type="solid")
    fills = {
        "Matched": PatternFill(start_color=ROW_MATCHED, end_color=ROW_MATCHED, fill_type="solid"),
        "Not in Statement": PatternFill(start_color=ROW_NOT_IN_STATEMENT, end_color=ROW_NOT_IN_STATEMENT, fill_type="solid"),
        "Not in Karibu": PatternFill(start_color=ROW_NOT_IN_KARIBU, end_color=ROW_NOT_IN_KARIBU, fill_type="solid"),
    }
    return fills.get(status)


def _status_font(status: str) -> Font:
    palette = {
        "Matched": BSR_DARK_GREEN,
        "Not in Statement": "C0392B",
        "Not in Karibu": "B7791A",
    }
    return Font(name="Arial", size=9, bold=True, color=palette.get(status, "000000"))


def _confidence_font(conf_str: str) -> Font:
    if not conf_str or conf_str == "—":
        return Font(name="Arial", size=9)
    try:
        val = int(conf_str.replace("%", ""))
    except ValueError:
        return Font(name="Arial", size=9)
    if val >= 90:
        c = BSR_DARK_GREEN
    elif val >= 70:
        c = "2471A3"
    elif val >= 50:
        c = "B7791A"
    else:
        c = "C0392B"
    return Font(name="Arial", size=9, bold=True, color=c)


def _auto_width(ws, max_width: int = 32) -> None:
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        widest = 0
        for cell in col[:60]:
            if cell.value is not None:
                widest = max(widest, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(widest + 2, max_width)


# ---------------------------------------------------------------------------
# Dashboard composer
# ---------------------------------------------------------------------------

def build_dashboard_lines(
    karibu_out: pd.DataFrame,
    stmt_out: pd.DataFrame,
    *,
    account_name: str,
    year: int,
    match_outflows: bool,
) -> list[str]:
    """Render dashboard text lines. Bidirectional accounts get DR/CR splits."""
    now = datetime.now().strftime("%d %b %Y")
    lines: list[str] = [
        f"BSR {account_name} Reconciliation Dashboard",
        f"Year: {year}     Generated: {now}",
        "",
    ]

    k_status = karibu_out["Status"] if "Status" in karibu_out.columns else pd.Series([], dtype=str)
    s_status = stmt_out["Status"] if "Status" in stmt_out.columns else pd.Series([], dtype=str)
    matched_mask = k_status == "Matched"
    nis_mask = k_status == "Not in Statement"
    nik_mask = s_status == "Not in Karibu"
    total_k = len(karibu_out)
    matched = int(matched_mask.sum())
    pct = (matched / total_k * 100) if total_k else 0.0

    lines.append("RECONCILIATION RESULTS")
    lines.append(f"  Matched:                   {matched} rows  ({pct:.1f}%)")
    lines.append(f"  Not in Statement:          {int(nis_mask.sum())} rows")
    lines.append(f"  Not in Karibu:             {int(nik_mask.sum())} rows")
    lines.append("")

    if match_outflows:
        # Split by Karibu side (DR=inflow, CR=outflow) and statement Direction.
        dr_mask = pd.to_numeric(karibu_out.get("DR (UGX)", 0), errors="coerce").fillna(0) > 0
        cr_mask = pd.to_numeric(karibu_out.get("CR (UGX)", 0), errors="coerce").fillna(0) > 0
        if "Direction" in stmt_out.columns and len(stmt_out):
            stmt_dir = stmt_out["Direction"].astype(str).str.upper()
        else:
            stmt_dir = pd.Series([], dtype=str)
        s_in = stmt_dir == "IN"
        s_out = stmt_dir == "OUT"
        lines.append("BIDIRECTIONAL BREAKDOWN")
        lines.append(f"  Karibu DR matched:         {int((matched_mask & dr_mask).sum())}")
        lines.append(f"  Karibu DR unmatched:       {int((nis_mask & dr_mask).sum())}")
        lines.append(f"  Karibu CR matched:         {int((matched_mask & cr_mask).sum())}")
        lines.append(f"  Karibu CR unmatched:       {int((nis_mask & cr_mask).sum())}")
        lines.append(f"  Statement IN unmatched:    {int((nik_mask & s_in).sum())}")
        lines.append(f"  Statement OUT unmatched:   {int((nik_mask & s_out).sum())}")
        lines.append("")

    # Audit flag totals
    flag_counts = _flag_counts(karibu_out, stmt_out)
    lines.append("AUDIT FLAGS SUMMARY")
    if flag_counts:
        for flag, count in sorted(flag_counts.items()):
            lines.append(f"  {flag}: {count} occurrences")
    else:
        lines.append("  No audit flags raised")
    return lines


def _flag_counts(karibu_out: pd.DataFrame, stmt_out: pd.DataFrame) -> dict[str, int]:
    counts: dict[str, int] = {}
    for df in (karibu_out, stmt_out):
        if "Audit Flag" not in df.columns:
            continue
        for v in df["Audit Flag"].dropna():
            for f in str(v).split(","):
                f = f.strip()
                if f:
                    counts[f] = counts.get(f, 0) + 1
    return counts


# ---------------------------------------------------------------------------
# Comment preservation (carried over from legacy core/reconciler)
# ---------------------------------------------------------------------------

def load_existing_comments(recon_path: Path) -> dict:
    """Pull previously-saved Comments out of an existing reconciliation file.

    Returned shape: `{"karibu": {key: comment}, "stmt": {txid: comment}}`.
    Keys are `f"{Date}|{Narration}|{DR (UGX)}"` so a rerun matches user
    annotations back to the same logical row even if positional order
    shifts. Best-effort — failures return empty dicts rather than raising.
    """
    comments = {"karibu": {}, "stmt": {}}
    if not recon_path.exists():
        return comments
    try:
        wb = load_workbook(recon_path, data_only=True)
    except Exception:
        return comments
    try:
        if "Karibu Report" in wb.sheetnames:
            ws = wb["Karibu Report"]
            headers = [cell.value for cell in ws[1]]
            try:
                comment_idx = headers.index("Comments")
                date_idx = headers.index("Date")
                narr_idx = headers.index("Narration")
                dr_idx = headers.index("DR (UGX)") if "DR (UGX)" in headers else -1
            except ValueError:
                comment_idx = -1
            if comment_idx >= 0:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if comment_idx >= len(row):
                        continue
                    cval = row[comment_idx]
                    if not cval:
                        continue
                    key = "|".join([
                        str(row[date_idx]) if date_idx < len(row) else "",
                        str(row[narr_idx]) if narr_idx < len(row) else "",
                        str(row[dr_idx]) if 0 <= dr_idx < len(row) else "",
                    ])
                    comments["karibu"][key] = cval

        for stmt_sheet in ("Statement", "Merchant Statement"):
            if stmt_sheet in wb.sheetnames:
                ws = wb[stmt_sheet]
                headers = [cell.value for cell in ws[1]]
                try:
                    comment_idx = headers.index("Comments")
                    txid_idx = headers.index("Transaction ID")
                except ValueError:
                    continue
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if comment_idx >= len(row) or txid_idx >= len(row):
                        continue
                    cval, txid = row[comment_idx], row[txid_idx]
                    if cval and txid:
                        comments["stmt"][str(txid)] = cval
                break
    finally:
        wb.close()
    return comments


def restore_comments(
    karibu_out: pd.DataFrame, stmt_out: pd.DataFrame, comments: dict
) -> None:
    if comments.get("karibu"):
        for idx in karibu_out.index:
            key = "|".join([
                str(karibu_out.at[idx, "Date"]) if "Date" in karibu_out.columns else "",
                str(karibu_out.at[idx, "Narration"]) if "Narration" in karibu_out.columns else "",
                str(karibu_out.at[idx, "DR (UGX)"]) if "DR (UGX)" in karibu_out.columns else "",
            ])
            if key in comments["karibu"]:
                karibu_out.at[idx, "Comments"] = comments["karibu"][key]
    if comments.get("stmt") and "Transaction ID" in stmt_out.columns:
        for idx in stmt_out.index:
            txid = str(stmt_out.at[idx, "Transaction ID"])
            if txid in comments["stmt"]:
                stmt_out.at[idx, "Comments"] = comments["stmt"][txid]
