"""Read consolidated Karibu / statement workbooks back into DataFrames.

The Phase-2 consolidator writes per-account, per-year workbooks under
`Statements/{Account}/`:
  - `{Account} Karibu Ledger - {YYYY}.xlsx`     (sheets: Jan..Dec [+ Unparseable])
  - `{Account} Transactions - {YYYY}.xlsx`      (sheets: Jan..Dec [+ Unparseable])

Phase-3 reconciliation reads all monthly sheets back, drops the
`Unparseable` review sheet (those rows have `date=None` so they can't
match anything anyway), and returns a single DataFrame ready for the
matching engine.

Column shapes (set by the consolidator — see consolidator._columns_for):

  Karibu:
    Date, Narration, Direction, Amount (UGX), DR, CR, Balance,
    Source File, Audit Flag

  Statement:
    Date, Transaction ID, Direction, Counterparty, Transaction Type,
    Amount (UGX), Source File, Audit Flag
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


_UNPARSEABLE_SHEET = "Unparseable"


class ConsolidatedFileNotFound(FileNotFoundError):
    """Raised when reconciler expects a consolidated workbook but can't find it."""


def _list_data_sheets(path: Path) -> list[str]:
    """Return month sheets, excluding the `Unparseable` review sheet."""
    wb = load_workbook(path, read_only=True)
    try:
        return [s for s in wb.sheetnames if s != _UNPARSEABLE_SHEET]
    finally:
        wb.close()


def count_unparseable(path: Path) -> int:
    """Return the row count on the `Unparseable` sheet, or 0 if absent."""
    if not path.exists():
        return 0
    wb = load_workbook(path, read_only=True)
    try:
        if _UNPARSEABLE_SHEET not in wb.sheetnames:
            return 0
        ws = wb[_UNPARSEABLE_SHEET]
        # Header on row 1; data rows from row 2.
        return max(ws.max_row - 1, 0)
    finally:
        wb.close()


def load_consolidated_karibu(path: Path) -> pd.DataFrame:
    """Concatenate every monthly Karibu sheet into one DataFrame.

    Date is coerced to pandas datetime; DR/CR/Amount columns to numeric.
    Empty workbook → empty DataFrame with the expected columns.
    """
    if not path.exists():
        raise ConsolidatedFileNotFound(path)
    sheets = _list_data_sheets(path)
    frames: list[pd.DataFrame] = []
    for sheet in sheets:
        df = pd.read_excel(path, sheet_name=sheet)
        if not df.empty:
            frames.append(df)
    if not frames:
        return pd.DataFrame(columns=[
            "Date", "Narration", "Direction", "Amount (UGX)",
            "DR", "CR", "Balance", "Source File", "Audit Flag",
        ])
    df = pd.concat(frames, ignore_index=True)
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    for col in ("DR", "CR", "Amount (UGX)"):
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    return df


def load_consolidated_statement(path: Path) -> pd.DataFrame:
    """Concatenate every monthly statement sheet into one DataFrame."""
    if not path.exists():
        raise ConsolidatedFileNotFound(path)
    sheets = _list_data_sheets(path)
    frames: list[pd.DataFrame] = []
    for sheet in sheets:
        df = pd.read_excel(path, sheet_name=sheet)
        if not df.empty:
            frames.append(df)
    if not frames:
        return pd.DataFrame(columns=[
            "Date", "Transaction ID", "Direction", "Counterparty",
            "Transaction Type", "Amount (UGX)", "Source File", "Audit Flag",
        ])
    df = pd.concat(frames, ignore_index=True)
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    if "Amount (UGX)" in df.columns:
        df["Amount (UGX)"] = pd.to_numeric(df["Amount (UGX)"], errors="coerce").fillna(0)
    if "Transaction ID" in df.columns:
        df["Transaction ID"] = df["Transaction ID"].astype(str).fillna("")
    return df
