"""End-to-end reconciler smoke tests on the Petty Cash UGX live samples.

These tests stage the MoMo + Karibu samples into a tmp directory that
looks like the post-Phase-2 XDG layout, run the consolidator to produce
yearly workbooks, then run `reconcile_account()` and verify the output
shape + key invariants (Karibu Report / Statement / Dashboard sheets,
BSR column layout, PETTY_CASH_NO_STATEMENT_EXPECTED flag presence,
UNMATCHED_HIGH_VALUE suppression on Karibu-only rows).
"""

from __future__ import annotations

import shutil
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from config import AccountConfig
from consolidator import consolidate_account
from reconciler import reconcile_account


def _petty_cash_account() -> AccountConfig:
    return AccountConfig(
        name="Petty Cash UGX",
        karibu_account="PC - Petty Cash UGX",
        statement_parser="momo_agent_xlsx",
        karibu_parser="karibu_ledger_csv",
        matching={"date_window_days": 2, "lumpsum_window_days": 2,
                  "amount_tolerance_ugx": 0.5},
        karibu_only_is_normal=True,
        match_outflows=True,
    )


def _stage(base: Path, momo_xlsx: Path, karibu_csv: Path) -> None:
    tx = base / "Transactions" / "Petty Cash UGX"
    kr = base / "Reports" / "Karibu" / "Petty Cash UGX"
    tx.mkdir(parents=True, exist_ok=True)
    kr.mkdir(parents=True, exist_ok=True)
    shutil.copy(momo_xlsx, tx / momo_xlsx.name)
    shutil.copy(karibu_csv, kr / karibu_csv.name)


def test_petty_cash_recon_produces_three_sheet_workbook(
    tmp_path: Path, momo_agent_xlsx, karibu_petty_cash_csv,
):
    _stage(tmp_path, momo_agent_xlsx, karibu_petty_cash_csv)
    account = _petty_cash_account()
    consolidate_account(account, tmp_path)

    result = reconcile_account(account, tmp_path, year=2026)
    assert result.output_path.exists()
    assert result.output_path.name == "Petty Cash UGX Reconciliation - 2026.xlsx"

    wb = load_workbook(result.output_path, read_only=True)
    assert wb.sheetnames == ["Karibu Report", "Statement", "Dashboard"]
    wb.close()


def test_petty_cash_recon_flags_unmatched_karibu_as_expected(
    tmp_path: Path, momo_agent_xlsx, karibu_petty_cash_csv,
):
    """Every Karibu 'Not in Statement' row must carry the soft flag and
    none of the hard-escalation flags."""
    _stage(tmp_path, momo_agent_xlsx, karibu_petty_cash_csv)
    account = _petty_cash_account()
    consolidate_account(account, tmp_path)

    result = reconcile_account(account, tmp_path, year=2026)
    wb = load_workbook(result.output_path)
    ws = wb["Karibu Report"]
    headers = [c.value for c in ws[1]]
    status_idx = headers.index("Status")
    flag_idx = headers.index("Audit Flag")
    soft_seen = False
    for row in ws.iter_rows(min_row=2, values_only=True):
        status = row[status_idx]
        flag = str(row[flag_idx] or "")
        if status == "Not in Statement":
            soft_seen = soft_seen or ("PETTY_CASH_NO_STATEMENT_EXPECTED" in flag)
            assert "UNMATCHED_HIGH_VALUE" not in flag
            assert "LARGE_SINGLE_PAYMENT" not in flag
            assert "DATE_GAP" not in flag
    wb.close()
    # On live data we should see at least some unmatched rows getting the flag.
    assert soft_seen, "No PETTY_CASH_NO_STATEMENT_EXPECTED flag observed"


def test_petty_cash_recon_result_carries_direction_splits(
    tmp_path: Path, momo_agent_xlsx, karibu_petty_cash_csv,
):
    _stage(tmp_path, momo_agent_xlsx, karibu_petty_cash_csv)
    account = _petty_cash_account()
    consolidate_account(account, tmp_path)
    result = reconcile_account(account, tmp_path, year=2026)
    # match_outflows=True populates the *_in / *_out fields, and they
    # should sum to the totals.
    assert result.matched_in + result.matched_out == result.matched
    assert (result.not_in_statement_in + result.not_in_statement_out
            == result.not_in_statement)
    assert (result.not_in_karibu_in + result.not_in_karibu_out
            == result.not_in_karibu)
