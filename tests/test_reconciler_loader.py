"""Tests for the reconciler loaders.

These tests build a tiny synthetic consolidated workbook on disk, then
read it back via the loader. We can't reuse the consolidator fixtures
here — that would couple two modules — so the test writes its own
month-sheet layout.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from reconciler.loader import (
    ConsolidatedFileNotFound,
    count_unparseable,
    load_consolidated_karibu,
    load_consolidated_statement,
)


def _write_karibu(path: Path, sheets: dict[str, list[list]]) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(sheet_name)
        headers = [
            "Date", "Narration", "Direction", "Amount (UGX)",
            "DR", "CR", "Balance", "Source File", "Audit Flag",
        ]
        ws.append(headers)
        for row in rows:
            ws.append(row)
    wb.save(path)


def _write_stmt(path: Path, sheets: dict[str, list[list]]) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(sheet_name)
        headers = [
            "Date", "Transaction ID", "Direction", "Counterparty",
            "Transaction Type", "Amount (UGX)", "Source File", "Audit Flag",
        ]
        ws.append(headers)
        for row in rows:
            ws.append(row)
    wb.save(path)


def test_loader_concatenates_monthly_sheets(tmp_path: Path):
    p = tmp_path / "k.xlsx"
    _write_karibu(p, {
        "Jan": [["2026-01-15", "narr1", "IN", 1000, 1000, 0, "", "x", ""]],
        "Feb": [
            ["2026-02-10", "narr2", "IN", 2000, 2000, 0, "", "x", ""],
            ["2026-02-12", "narr3", "OUT", 500, 0, 500, "", "x", ""],
        ],
    })
    df = load_consolidated_karibu(p)
    assert len(df) == 3
    assert df["DR"].sum() == 3000
    assert df["CR"].sum() == 500


def test_loader_skips_unparseable_sheet(tmp_path: Path):
    """The Unparseable review sheet must NOT enter the matching pipeline —
    those rows have date=None and would always be Not in Statement noise."""
    p = tmp_path / "k.xlsx"
    _write_karibu(p, {
        "Jan": [["2026-01-15", "good", "IN", 1000, 1000, 0, "", "x", ""]],
        "Unparseable": [["NOT-A-DATE", "bad", "IN", 9999, 9999, 0, "", "x", "UNPARSEABLE_DATE"]],
    })
    df = load_consolidated_karibu(p)
    assert len(df) == 1
    assert df.iloc[0]["Narration"] == "good"


def test_loader_count_unparseable_returns_row_count(tmp_path: Path):
    p = tmp_path / "k.xlsx"
    _write_karibu(p, {
        "Jan": [["2026-01-15", "good", "IN", 1000, 1000, 0, "", "x", ""]],
        "Unparseable": [
            ["X", "bad1", "IN", 9999, 9999, 0, "", "x", "UNPARSEABLE_DATE"],
            ["X", "bad2", "IN", 9999, 9999, 0, "", "x", "UNPARSEABLE_DATE"],
        ],
    })
    assert count_unparseable(p) == 2


def test_loader_empty_workbook(tmp_path: Path):
    """A workbook with one header-only sheet must still load cleanly —
    the consolidator can produce this for accounts with zero records."""
    p = tmp_path / "k.xlsx"
    _write_karibu(p, {"Jan": []})
    df = load_consolidated_karibu(p)
    assert df.empty
    assert "DR" in df.columns
    assert "CR" in df.columns


def test_loader_statement_concats(tmp_path: Path):
    p = tmp_path / "s.xlsx"
    _write_stmt(p, {
        "Mar": [["2026-03-15 10:00", "TXN-1", "IN", "Alice", "CASH_IN", 5000, "x", ""]],
        "Apr": [["2026-04-15 09:00", "TXN-2", "OUT", "Bob", "TRANSFER", 1000, "x", ""]],
    })
    df = load_consolidated_statement(p)
    assert len(df) == 2
    assert df["Amount (UGX)"].sum() == 6000


def test_loader_missing_file_raises(tmp_path: Path):
    with pytest.raises(ConsolidatedFileNotFound):
        load_consolidated_karibu(tmp_path / "absent.xlsx")
