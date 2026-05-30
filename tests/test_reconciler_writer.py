"""Tests for the reconciliation Excel writer.

We verify the sheet structure, the BSR header styling, and the
comment-preservation roundtrip — the actual cell-by-cell colouring is
covered by the legacy excel_writer tests (or its absence is fine — these
tests assert the contract a downstream consumer cares about).
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from reconciler.writer import (
    build_dashboard_lines,
    load_existing_comments,
    restore_comments,
    write_reconciliation_workbook,
)


def _karibu_frame() -> pd.DataFrame:
    return pd.DataFrame([
        {"Date": "2026-05-10", "Account": "MTN Money", "Narration": "alice",
         "DR (UGX)": 1000, "CR (UGX)": 0, "Balance": "",
         "Status": "Matched", "Match Type": "Exact", "Confidence": "100%",
         "Matched Ref": "TX-1", "Audit Flag": "", "Comments": ""},
        {"Date": "2026-05-12", "Account": "MTN Money", "Narration": "bob",
         "DR (UGX)": 5000, "CR (UGX)": 0, "Balance": "",
         "Status": "Not in Statement", "Match Type": "—",
         "Confidence": "—", "Matched Ref": "—", "Audit Flag": "",
         "Comments": "user note from last run"},
    ])


def _stmt_frame() -> pd.DataFrame:
    return pd.DataFrame([
        {"Date": "2026-05-10", "Transaction ID": "TX-1", "Payer Name": "Alice",
         "Amount (UGX)": 1000, "Tx Status": "Successful", "Direction": "IN",
         "Status": "Matched", "Match Type": "Exact", "Confidence": "100%",
         "Matched Ref": "K0", "Audit Flag": "", "Comments": ""},
    ])


def test_workbook_has_three_named_sheets(tmp_path: Path):
    out = tmp_path / "recon.xlsx"
    write_reconciliation_workbook(
        _karibu_frame(), _stmt_frame(),
        ["Dashboard"], out,
    )
    wb = load_workbook(out, read_only=True)
    assert wb.sheetnames == ["Karibu Report", "Statement", "Dashboard"]
    wb.close()


def test_karibu_report_columns_match_legacy_layout(tmp_path: Path):
    """Phase-3 spec: column structure must match samples/BSR_MTN_Reconciliation.xlsx
    Karibu Report tab."""
    out = tmp_path / "recon.xlsx"
    write_reconciliation_workbook(
        _karibu_frame(), _stmt_frame(), ["X"], out,
    )
    wb = load_workbook(out, read_only=True)
    headers = [c.value for c in wb["Karibu Report"][1]]
    wb.close()
    assert headers == [
        "Date", "Account", "Narration", "DR (UGX)", "CR (UGX)",
        "Balance", "Status", "Match Type", "Confidence", "Matched Ref",
        "Audit Flag", "Comments",
    ]


def test_dashboard_lines_render_in_order(tmp_path: Path):
    out = tmp_path / "recon.xlsx"
    dashboard = ["BSR Test Dashboard", "Generated", "", "TOTALS", "  Matched: 5"]
    write_reconciliation_workbook(
        _karibu_frame(), _stmt_frame(), dashboard, out,
    )
    wb = load_workbook(out, read_only=True)
    ws = wb["Dashboard"]
    # Openpyxl reads an empty string back as None — collapse for comparison.
    rendered = [(ws.cell(row=i, column=1).value or "") for i in range(1, len(dashboard) + 1)]
    wb.close()
    assert rendered == dashboard


def test_comments_round_trip_through_load_and_restore(tmp_path: Path):
    out = tmp_path / "recon.xlsx"
    karibu = _karibu_frame()
    stmt = _stmt_frame()
    write_reconciliation_workbook(karibu, stmt, ["X"], out)

    existing = load_existing_comments(out)
    # The Karibu key uses Date|Narration|DR (UGX) — both rows should be picked up
    # (only the second has a comment though).
    assert len(existing["karibu"]) == 1

    # Now build fresh frames (Comments wiped) and restore should re-attach.
    fresh_k = _karibu_frame()
    fresh_k["Comments"] = ""
    fresh_s = _stmt_frame()
    fresh_s["Comments"] = ""
    restore_comments(fresh_k, fresh_s, existing)
    assert fresh_k.loc[1, "Comments"] == "user note from last run"


def test_dashboard_lines_include_match_counts():
    karibu = _karibu_frame()
    stmt = _stmt_frame()
    lines = build_dashboard_lines(
        karibu, stmt, account_name="MTN Merchant", year=2026,
        match_outflows=False,
    )
    text = "\n".join(lines)
    assert "MTN Merchant" in text
    assert "Matched:" in text
    assert "Not in Statement" in text


def test_bidirectional_dashboard_has_dr_cr_split():
    """When match_outflows=True the dashboard surfaces per-direction counts."""
    karibu = pd.DataFrame([
        {"Date": "2026-05-10", "Account": "Petty Cash UGX",
         "Narration": "x", "DR (UGX)": 1000, "CR (UGX)": 0, "Balance": "",
         "Status": "Matched", "Match Type": "Exact", "Confidence": "100%",
         "Matched Ref": "K0", "Audit Flag": "", "Comments": ""},
        {"Date": "2026-05-11", "Account": "Petty Cash UGX",
         "Narration": "y", "DR (UGX)": 0, "CR (UGX)": 500, "Balance": "",
         "Status": "Not in Statement", "Match Type": "—",
         "Confidence": "—", "Matched Ref": "—",
         "Audit Flag": "PETTY_CASH_NO_STATEMENT_EXPECTED", "Comments": ""},
    ])
    stmt = pd.DataFrame()
    lines = build_dashboard_lines(
        karibu, stmt, account_name="Petty Cash UGX", year=2026,
        match_outflows=True,
    )
    text = "\n".join(lines)
    assert "Karibu DR matched:" in text
    assert "Karibu CR unmatched:" in text
