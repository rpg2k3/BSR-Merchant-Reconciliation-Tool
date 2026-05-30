"""MTN parity test: the new reconciler must match `core/reconciler.py`
results on the same MTN-style inputs to within ±1.

This protects MTN Merchant + Airtel Merchant users from an unintended
behaviour change when Phase 3 cuts over to the new package. The test
plants a small but representative synthetic dataset exercising several
of the 7 passes, runs both engines through their public entry points,
and compares the matched-Karibu counts.
"""

from __future__ import annotations

import shutil
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from config import AccountConfig
from consolidator import consolidate_account
from core.reconciler import reconcile as legacy_reconcile
from reconciler import reconcile_account


# ---------------------------------------------------------------------------
# Synthetic-input helpers
# ---------------------------------------------------------------------------

def _mtn_account() -> AccountConfig:
    return AccountConfig(
        name="MTN Merchant",
        karibu_account="MTN Money",
        statement_parser="mtn_merchant_csv",
        karibu_parser="karibu_ledger_csv",
        matching={"date_window_days": 2, "lumpsum_window_days": 0,
                  "amount_tolerance_ugx": 0.5},
        legacy_folder="MTN",
        karibu_only_is_normal=False,
        match_outflows=False,
    )


def _write_mtn_portal_csv(path: Path, rows: list[dict]) -> None:
    """Write a CSV that looks like an MTN merchant portal export."""
    df = pd.DataFrame(rows)
    df.to_csv(path, index=False)


def _write_legacy_mtn_xlsx(path: Path, rows: list[dict]) -> None:
    """Write the legacy 'BSR_MTN_Merchant_Transactions.xlsx' that
    core/reconciler.load_mtn_statement() expects.

    Sheet 'MTN Transactions', row 1 banner, row 2 headers, data row 3+.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "MTN Transactions"
    ws.cell(row=1, column=1, value="BSR MTN Merchant Statement — Synthetic Test")
    headers = list(rows[0].keys())
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=2, column=col_idx, value=h)
    for r_idx, row in enumerate(rows, 3):
        for c_idx, h in enumerate(headers, 1):
            ws.cell(row=r_idx, column=c_idx, value=row[h])
    wb.save(path)


def _write_karibu_csv(path: Path, rows: list[dict]) -> None:
    """Write a CSV in the Karibu ledger export format.

    The format is `skiprows=2`, so the first two rows are noise (title +
    blank), then a header row, then data rows.
    """
    df = pd.DataFrame(rows)
    with path.open("w", encoding="utf-8") as f:
        f.write("Karibu HMS Ledger Export\n\n")
    df.to_csv(path, mode="a", index=False)


# ---------------------------------------------------------------------------
# The synthetic dataset
# ---------------------------------------------------------------------------

# Eight Karibu DR rows; eight statement IN rows. Plus a few decoys so we
# exercise more than just same-day exact match.
KARIBU_ROWS = [
    # Exact same-day matches
    {"Date": "2026-04-01", "Account": "MTN Money", "Narration": "alpha",
     "DR": "100000", "CR": "0", "Balance": "100000"},
    {"Date": "2026-04-02", "Account": "MTN Money", "Narration": "beta",
     "DR": "200000", "CR": "0", "Balance": "300000"},
    {"Date": "2026-04-03", "Account": "MTN Money", "Narration": "gamma",
     "DR": "50000", "CR": "0", "Balance": "350000"},
    # ±1 day drift
    {"Date": "2026-04-04", "Account": "MTN Money", "Narration": "delta",
     "DR": "75000", "CR": "0", "Balance": "425000"},
    # Lumpsum K→S (one Karibu = two stmt rows on same day)
    {"Date": "2026-04-05", "Account": "MTN Money", "Narration": "epsilon",
     "DR": "30000", "CR": "0", "Balance": "455000"},
    # Amount-only (date far apart)
    {"Date": "2026-04-06", "Account": "MTN Money", "Narration": "zeta",
     "DR": "999", "CR": "0", "Balance": "455999"},
    # No matching statement row — leave as Not in Statement on both engines.
    {"Date": "2026-04-07", "Account": "MTN Money", "Narration": "orphan",
     "DR": "55", "CR": "0", "Balance": "456054"},
    # Karibu CR row — both engines must ignore on the matching path.
    {"Date": "2026-04-08", "Account": "MTN Money", "Narration": "outflow",
     "DR": "0", "CR": "12345", "Balance": "443709"},
]

STMT_ROWS = [
    {"Date": "2026-04-01 10:00:00", "Id": "TX-A", "Amount": 100000,
     "From name": "Customer 1", "Status": "Successful"},
    {"Date": "2026-04-02 10:00:00", "Id": "TX-B", "Amount": 200000,
     "From name": "Customer 2", "Status": "Successful"},
    {"Date": "2026-04-03 10:00:00", "Id": "TX-C", "Amount": 50000,
     "From name": "Customer 3", "Status": "Successful"},
    # one-day drift (Karibu has 04-04, stmt has 04-05) → ±1 pass.
    {"Date": "2026-04-05 10:00:00", "Id": "TX-D", "Amount": 75000,
     "From name": "Customer 4", "Status": "Successful"},
    # lumpsum: 20000 + 10000 = 30000 to match epsilon on 2026-04-05.
    {"Date": "2026-04-05 11:00:00", "Id": "TX-E1", "Amount": 20000,
     "From name": "Customer 5", "Status": "Successful"},
    {"Date": "2026-04-05 12:00:00", "Id": "TX-E2", "Amount": 10000,
     "From name": "Customer 6", "Status": "Successful"},
    # amount-only: 999 on a totally different date.
    {"Date": "2026-05-20 10:00:00", "Id": "TX-F", "Amount": 999,
     "From name": "Customer 7", "Status": "Successful"},
    # Statement row with no Karibu counterpart.
    {"Date": "2026-04-10 10:00:00", "Id": "TX-G", "Amount": 9999,
     "From name": "Customer 8", "Status": "Successful"},
]


# ---------------------------------------------------------------------------
# Pipeline drivers
# ---------------------------------------------------------------------------

def _run_legacy(base: Path) -> int:
    """Stage legacy layout, run core/reconciler, return matched count."""
    statements = base / "Statements"
    reports = base / "Reports" / "Karibu" / "MTN"
    recon = base / "Reconciliation"
    backups = base / "Backups"
    for d in (statements, reports, recon, backups):
        d.mkdir(parents=True, exist_ok=True)
    _write_legacy_mtn_xlsx(statements / "BSR_MTN_Merchant_Transactions.xlsx", STMT_ROWS)
    _write_karibu_csv(reports / "karibu_2026.csv", KARIBU_ROWS)

    result = legacy_reconcile("MTN", base, config={
        "date_tolerance_days": 2,
        "high_value_threshold": 500_000,
        "large_payment_threshold": 1_000_000,
    })
    return int(result.get("matched", -1))


def _run_new(base: Path) -> int:
    """Stage new layout, run consolidator + new reconciler, return matched count."""
    tx_dir = base / "Transactions" / "MTN Merchant"
    karibu_dir = base / "Reports" / "Karibu" / "MTN Merchant"
    tx_dir.mkdir(parents=True, exist_ok=True)
    karibu_dir.mkdir(parents=True, exist_ok=True)
    _write_mtn_portal_csv(tx_dir / "mtn_2026.csv", STMT_ROWS)
    _write_karibu_csv(karibu_dir / "karibu_2026.csv", KARIBU_ROWS)

    account = _mtn_account()
    consolidate_account(account, base)
    result = reconcile_account(account, base, year=2026)
    return result.matched


# ---------------------------------------------------------------------------
# The parity test
# ---------------------------------------------------------------------------

def test_mtn_matched_count_within_one_of_legacy(tmp_path: Path):
    """Both engines must agree on matched Karibu count to within ±1.

    A bigger gap means the new engine silently changed semantics; per the
    Phase-3 spec, that requires investigation before the cutover.
    """
    legacy_dir = tmp_path / "legacy"
    new_dir = tmp_path / "new"
    legacy_matched = _run_legacy(legacy_dir)
    new_matched = _run_new(new_dir)

    assert legacy_matched >= 0, "legacy reconciler errored"
    delta = abs(new_matched - legacy_matched)
    assert delta <= 1, (
        f"matched-count drift: legacy={legacy_matched} new={new_matched} "
        f"(±1 allowed, observed {delta})"
    )


def test_new_reconciler_match_outflows_off_omits_cr_rows(tmp_path: Path):
    """With match_outflows=False, Karibu CR rows must NOT appear in the
    Karibu Report sheet. This mirrors the legacy MTN/Airtel behaviour."""
    new_dir = tmp_path / "new"
    _run_new(new_dir)
    out = new_dir / "Reconciliation" / "MTN Merchant" / "MTN Merchant Reconciliation - 2026.xlsx"
    assert out.exists()
    wb = load_workbook(out, read_only=True)
    ws = wb["Karibu Report"]
    headers = [c.value for c in ws[1]]
    narration_idx = headers.index("Narration")
    narrations = [row[narration_idx] for row in ws.iter_rows(min_row=2, values_only=True)]
    wb.close()
    assert "outflow" not in narrations, (
        "Karibu CR row leaked into Karibu Report with match_outflows=False"
    )
