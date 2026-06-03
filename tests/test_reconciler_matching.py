"""Tests for the 7-pass matching engine.

Most run on tiny hand-built DataFrames so behaviour can be reasoned about
row-by-row. The end-to-end `match_outflows=False` test at the bottom was
rescued from the deleted `test_reconciler_mtn_parity.py` (the parity check it
sat beside compared against the now-removed legacy `core.reconciler`; this one
only drives the new pipeline and stays a live correctness guarantee).
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from config import AccountConfig
from consolidator import consolidate_account
from reconciler import reconcile_account
from reconciler.matching import (
    _greedy_subset_sum,
    run_matching,
)
from reconciler.types import ReconKnobs


def _karibu(rows: list[dict]) -> pd.DataFrame:
    return pd.DataFrame(rows)


def _stmt(rows: list[dict]) -> pd.DataFrame:
    return pd.DataFrame(rows)


def test_pass1_same_day_exact_yields_100pct():
    karibu = _karibu([{"Date": datetime(2026, 5, 10), "DR": 1000, "CR": 0}])
    stmt = _stmt([{"Date": datetime(2026, 5, 10),
                   "Amount (UGX)": 1000, "Direction": "IN",
                   "Transaction ID": "TX-1"}])
    k_res, s_res = run_matching(karibu, stmt, ReconKnobs())
    assert k_res.iloc[0]["Status"] == "Matched"
    assert k_res.iloc[0]["Confidence"] == "100%"
    assert k_res.iloc[0]["Match Type"] == "Exact"
    assert s_res.iloc[0]["Matched Ref"] == "K0"


def test_pass2_one_day_drift_yields_90pct():
    karibu = _karibu([{"Date": datetime(2026, 5, 10), "DR": 1000, "CR": 0}])
    stmt = _stmt([{"Date": datetime(2026, 5, 11),
                   "Amount (UGX)": 1000, "Direction": "IN",
                   "Transaction ID": "TX-1"}])
    k_res, _ = run_matching(karibu, stmt, ReconKnobs())
    assert k_res.iloc[0]["Confidence"] == "90%"


def test_pass3_two_day_drift_yields_80pct():
    karibu = _karibu([{"Date": datetime(2026, 5, 10), "DR": 1000, "CR": 0}])
    stmt = _stmt([{"Date": datetime(2026, 5, 12),
                   "Amount (UGX)": 1000, "Direction": "IN",
                   "Transaction ID": "TX-1"}])
    k_res, _ = run_matching(karibu, stmt, ReconKnobs())
    assert k_res.iloc[0]["Confidence"] == "80%"


def test_lumpsum_k_to_s_combines_two_stmt_rows():
    """One Karibu DR of 3000 matches two same-day statement rows of 1000+2000."""
    karibu = _karibu([{"Date": datetime(2026, 5, 10), "DR": 3000, "CR": 0}])
    stmt = _stmt([
        {"Date": datetime(2026, 5, 10), "Amount (UGX)": 1000,
         "Direction": "IN", "Transaction ID": "A"},
        {"Date": datetime(2026, 5, 10), "Amount (UGX)": 2000,
         "Direction": "IN", "Transaction ID": "B"},
    ])
    k_res, s_res = run_matching(karibu, stmt, ReconKnobs())
    assert k_res.iloc[0]["Match Type"] == "Lumpsum"
    assert k_res.iloc[0]["Confidence"] == "60%"
    assert s_res["Status"].tolist() == ["Matched", "Matched"]


def test_lumpsum_s_to_k_respects_lumpsum_window_days():
    """One stmt row of 3000 across two Karibu DRs on different days inside lumpsum_window_days."""
    karibu = _karibu([
        {"Date": datetime(2026, 5, 10), "DR": 1000, "CR": 0},
        {"Date": datetime(2026, 5, 12), "DR": 2000, "CR": 0},
    ])
    stmt = _stmt([{"Date": datetime(2026, 5, 11), "Amount (UGX)": 3000,
                   "Direction": "IN", "Transaction ID": "BIG"}])
    # With lumpsum_window_days=0 (legacy MTN), the two Karibu rows are
    # outside the same-day window → no match. With lumpsum_window_days=2
    # (Petty Cash), the match is found.
    k0, s0 = run_matching(karibu, stmt, ReconKnobs(lumpsum_window_days=0))
    assert (k0["Status"] == "Not in Statement").all()

    k2, s2 = run_matching(karibu, stmt, ReconKnobs(lumpsum_window_days=2))
    assert (k2["Status"] == "Matched").all()
    assert s2.iloc[0]["Status"] == "Matched"


def test_pass7_amount_only_with_far_apart_dates():
    """Same amount, 30 days apart → amount-only match at 40%."""
    karibu = _karibu([{"Date": datetime(2026, 1, 1), "DR": 1000, "CR": 0}])
    stmt = _stmt([{"Date": datetime(2026, 5, 1), "Amount (UGX)": 1000,
                   "Direction": "IN", "Transaction ID": "X"}])
    k_res, _ = run_matching(karibu, stmt, ReconKnobs())
    assert k_res.iloc[0]["Confidence"] == "40%"
    assert k_res.iloc[0]["Match Type"] == "Amount Only"


def test_match_outflows_false_ignores_cr_side():
    """Legacy MTN/Airtel: CR rows never match against OUT statement rows."""
    karibu = _karibu([{"Date": datetime(2026, 5, 10), "DR": 0, "CR": 1000}])
    stmt = _stmt([{"Date": datetime(2026, 5, 10), "Amount (UGX)": 1000,
                   "Direction": "OUT", "Transaction ID": "X"}])
    k_res, s_res = run_matching(karibu, stmt, ReconKnobs(),
                                 match_outflows=False)
    # CR row passes through; matching engine never touched it.
    assert k_res.iloc[0]["Status"] == "Not in Statement"
    assert s_res.iloc[0]["Status"] == "Not in Karibu"


def test_match_outflows_true_matches_cr_with_out():
    """Petty Cash: Karibu CR matches statement OUT row."""
    karibu = _karibu([{"Date": datetime(2026, 5, 10), "DR": 0, "CR": 1000}])
    stmt = _stmt([{"Date": datetime(2026, 5, 10), "Amount (UGX)": 1000,
                   "Direction": "OUT", "Transaction ID": "X"}])
    k_res, s_res = run_matching(karibu, stmt, ReconKnobs(),
                                 match_outflows=True)
    assert k_res.iloc[0]["Status"] == "Matched"
    assert s_res.iloc[0]["Status"] == "Matched"


def test_match_outflows_true_keeps_dr_in_pipeline_too():
    """Bidirectional accounts must STILL match DR↔IN."""
    karibu = _karibu([
        {"Date": datetime(2026, 5, 10), "DR": 1000, "CR": 0},
        {"Date": datetime(2026, 5, 10), "DR": 0, "CR": 500},
    ])
    stmt = _stmt([
        {"Date": datetime(2026, 5, 10), "Amount (UGX)": 1000,
         "Direction": "IN", "Transaction ID": "IN-1"},
        {"Date": datetime(2026, 5, 10), "Amount (UGX)": 500,
         "Direction": "OUT", "Transaction ID": "OUT-1"},
    ])
    k_res, s_res = run_matching(karibu, stmt, ReconKnobs(), match_outflows=True)
    assert (k_res["Status"] == "Matched").all()
    assert (s_res["Status"] == "Matched").all()


def test_amount_tolerance_lets_half_ugx_drift_through():
    """0.5 UGX tolerance — legacy rounding semantics."""
    karibu = _karibu([{"Date": datetime(2026, 5, 10), "DR": 1000.3, "CR": 0}])
    stmt = _stmt([{"Date": datetime(2026, 5, 10),
                   "Amount (UGX)": 1000.0, "Direction": "IN",
                   "Transaction ID": "X"}])
    k_res, _ = run_matching(karibu, stmt, ReconKnobs(amount_tolerance_ugx=0.5))
    assert k_res.iloc[0]["Status"] == "Matched"


def test_locked_row_not_reclaimed_by_later_pass():
    """Once a Karibu row is matched at 100%, pass 7 doesn't steal it."""
    karibu = _karibu([{"Date": datetime(2026, 5, 10), "DR": 1000, "CR": 0}])
    stmt = _stmt([
        {"Date": datetime(2026, 5, 10), "Amount (UGX)": 1000,
         "Direction": "IN", "Transaction ID": "GOOD"},
        # A second statement row with the same amount on a far-away date —
        # would match in pass 7 if the locked row weren't honoured.
        {"Date": datetime(2026, 8, 1), "Amount (UGX)": 1000,
         "Direction": "IN", "Transaction ID": "BAD"},
    ])
    k_res, s_res = run_matching(karibu, stmt, ReconKnobs())
    assert k_res.iloc[0]["Confidence"] == "100%"
    assert s_res.iloc[0]["Status"] == "Matched"  # GOOD
    assert s_res.iloc[1]["Status"] == "Not in Karibu"  # BAD


# ---------- subset-sum helper ----------

def test_subset_sum_single_hit():
    res = _greedy_subset_sum([(0, 1000.0), (1, 500.0)], target=1000.0, tolerance=0.5)
    assert res == [(0, 1000.0)]


def test_subset_sum_two_member_combo():
    res = _greedy_subset_sum([(0, 700.0), (1, 300.0), (2, 50.0)], target=1000.0, tolerance=0.5)
    assert res is not None
    assert sum(amt for _, amt in res) == pytest.approx(1000.0)


def test_subset_sum_no_match_returns_none():
    res = _greedy_subset_sum([(0, 333.0), (1, 444.0)], target=1000.0, tolerance=0.5)
    assert res is None


def test_subset_sum_empty_candidates():
    assert _greedy_subset_sum([], target=1.0, tolerance=0.5) is None


# ---------------------------------------------------------------------------
# End-to-end: match_outflows=False keeps Karibu CR rows out of the report.
#
# Rescued from the deleted tests/test_reconciler_mtn_parity.py. Unlike the
# unit-level `test_match_outflows_false_ignores_cr_side` above (which calls
# `run_matching` directly), this drives the full consolidate -> reconcile
# pipeline and asserts the CR row never reaches the written `Karibu Report`
# sheet — a correctness guarantee for MTN/Airtel reconciliation.
# ---------------------------------------------------------------------------

def _mtn_account() -> AccountConfig:
    return AccountConfig(
        name="MTN Merchant",
        karibu_account="MTN Money",
        statement_parser="mtn_merchant_csv",
        karibu_parser="karibu_ledger_csv",
        matching={"date_window_days": 2, "lumpsum_window_days": 0,
                  "amount_tolerance_ugx": 0.5},
        legacy_folder="MTN",
        karibu_only_is_normal=False,
        match_outflows=False,
    )


# Seven DR rows exercising several passes, plus one CR row ("outflow") that
# must be ignored on the matching path and excluded from the Karibu Report.
_KARIBU_ROWS = [
    {"Date": "2026-04-01", "Account": "MTN Money", "Narration": "alpha",
     "DR": "100000", "CR": "0", "Balance": "100000"},
    {"Date": "2026-04-02", "Account": "MTN Money", "Narration": "beta",
     "DR": "200000", "CR": "0", "Balance": "300000"},
    {"Date": "2026-04-03", "Account": "MTN Money", "Narration": "gamma",
     "DR": "50000", "CR": "0", "Balance": "350000"},
    {"Date": "2026-04-04", "Account": "MTN Money", "Narration": "delta",
     "DR": "75000", "CR": "0", "Balance": "425000"},
    {"Date": "2026-04-05", "Account": "MTN Money", "Narration": "epsilon",
     "DR": "30000", "CR": "0", "Balance": "455000"},
    {"Date": "2026-04-06", "Account": "MTN Money", "Narration": "zeta",
     "DR": "999", "CR": "0", "Balance": "455999"},
    {"Date": "2026-04-07", "Account": "MTN Money", "Narration": "orphan",
     "DR": "55", "CR": "0", "Balance": "456054"},
    # Karibu CR row — must NOT appear in the Karibu Report with match_outflows=False.
    {"Date": "2026-04-08", "Account": "MTN Money", "Narration": "outflow",
     "DR": "0", "CR": "12345", "Balance": "443709"},
]

_STMT_ROWS = [
    {"Date": "2026-04-01 10:00:00", "Id": "TX-A", "Amount": 100000,
     "From name": "Customer 1", "Status": "Successful"},
    {"Date": "2026-04-02 10:00:00", "Id": "TX-B", "Amount": 200000,
     "From name": "Customer 2", "Status": "Successful"},
    {"Date": "2026-04-03 10:00:00", "Id": "TX-C", "Amount": 50000,
     "From name": "Customer 3", "Status": "Successful"},
    {"Date": "2026-04-05 10:00:00", "Id": "TX-D", "Amount": 75000,
     "From name": "Customer 4", "Status": "Successful"},
    {"Date": "2026-04-05 11:00:00", "Id": "TX-E1", "Amount": 20000,
     "From name": "Customer 5", "Status": "Successful"},
    {"Date": "2026-04-05 12:00:00", "Id": "TX-E2", "Amount": 10000,
     "From name": "Customer 6", "Status": "Successful"},
    {"Date": "2026-05-20 10:00:00", "Id": "TX-F", "Amount": 999,
     "From name": "Customer 7", "Status": "Successful"},
    {"Date": "2026-04-10 10:00:00", "Id": "TX-G", "Amount": 9999,
     "From name": "Customer 8", "Status": "Successful"},
]


def _write_mtn_portal_csv(path: Path, rows: list[dict]) -> None:
    """Write a CSV that looks like an MTN merchant portal export."""
    pd.DataFrame(rows).to_csv(path, index=False)


def _write_karibu_csv(path: Path, rows: list[dict]) -> None:
    """Write a Karibu ledger export CSV (two noise rows, then header+data —
    the `skiprows=2` convention the parser expects)."""
    with path.open("w", encoding="utf-8") as f:
        f.write("Karibu HMS Ledger Export\n\n")
    pd.DataFrame(rows).to_csv(path, mode="a", index=False)


def _run_new_mtn_pipeline(base: Path) -> None:
    """Stage synthetic inputs, then run consolidator + new reconciler."""
    tx_dir = base / "Transactions" / "MTN Merchant"
    karibu_dir = base / "Reports" / "Karibu" / "MTN Merchant"
    tx_dir.mkdir(parents=True, exist_ok=True)
    karibu_dir.mkdir(parents=True, exist_ok=True)
    _write_mtn_portal_csv(tx_dir / "mtn_2026.csv", _STMT_ROWS)
    _write_karibu_csv(karibu_dir / "karibu_2026.csv", _KARIBU_ROWS)

    account = _mtn_account()
    consolidate_account(account, base)
    reconcile_account(account, base, year=2026)


def test_new_reconciler_match_outflows_off_omits_cr_rows(tmp_path: Path):
    """With match_outflows=False, Karibu CR rows must NOT appear in the
    Karibu Report sheet. This mirrors the legacy MTN/Airtel behaviour."""
    new_dir = tmp_path / "new"
    _run_new_mtn_pipeline(new_dir)
    out = new_dir / "Reconciliation" / "MTN Merchant" / "MTN Merchant Reconciliation - 2026.xlsx"
    assert out.exists()
    wb = load_workbook(out, read_only=True)
    ws = wb["Karibu Report"]
    headers = [c.value for c in ws[1]]
    narration_idx = headers.index("Narration")
    narrations = [row[narration_idx] for row in ws.iter_rows(min_row=2, values_only=True)]
    wb.close()
    assert "outflow" not in narrations, (
        "Karibu CR row leaked into Karibu Report with match_outflows=False"
    )
