"""Phase 2 consolidator tests.

These tests exercise the consolidator end-to-end against synthetic
fixtures laid out under a tmp_path, plus the live MoMo + Karibu samples
from samples/ where convenient.
"""

from __future__ import annotations

import hashlib
import shutil
from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from config import AccountConfig
from consolidator import (
    ConsolidateResult,
    MONTH_ABBREV,
    _dedupe,
    _karibu_dedup_key,
    _statement_dedup_key,
    consolidate_account,
)
from parsers.types import AUDIT_UNPARSEABLE_DATE


def _petty_cash_account() -> AccountConfig:
    return AccountConfig(
        name="Petty Cash UGX",
        karibu_account="PC - Petty Cash UGX",
        statement_parser="momo_agent_xlsx",
        karibu_parser="karibu_ledger_csv",
        matching={"date_window_days": 2, "lumpsum_window_days": 2,
                  "amount_tolerance_ugx": 0},
        karibu_only_is_normal=True,
    )


def _stage_petty_cash_inputs(base: Path, momo_xlsx: Path, karibu_csv: Path) -> None:
    """Drop one MoMo xlsx and one Karibu CSV into the per-account input
    folders that the consolidator expects."""
    tx = base / "Transactions" / "Petty Cash UGX"
    kr = base / "Reports" / "Karibu" / "Petty Cash UGX"
    tx.mkdir(parents=True, exist_ok=True)
    kr.mkdir(parents=True, exist_ok=True)
    shutil.copy(momo_xlsx, tx / momo_xlsx.name)
    shutil.copy(karibu_csv, kr / karibu_csv.name)


def test_consolidator_produces_per_year_workbooks(
    tmp_path: Path, momo_agent_xlsx, karibu_petty_cash_csv
):
    _stage_petty_cash_inputs(tmp_path, momo_agent_xlsx, karibu_petty_cash_csv)

    result = consolidate_account(_petty_cash_account(), tmp_path)
    assert isinstance(result, ConsolidateResult)
    assert result.statement_records_unique == 354
    assert result.karibu_records_unique == 686

    out_dir = tmp_path / "Statements" / "Petty Cash UGX"
    # MoMo sample covers Mar–May 2026; Karibu sample covers Jan–May 2026.
    assert (out_dir / "Petty Cash UGX Transactions - 2026.xlsx").exists()
    assert (out_dir / "Petty Cash UGX Karibu Ledger - 2026.xlsx").exists()


def test_consolidator_monthly_sheets_have_expected_layout(
    tmp_path: Path, momo_agent_xlsx, karibu_petty_cash_csv
):
    _stage_petty_cash_inputs(tmp_path, momo_agent_xlsx, karibu_petty_cash_csv)
    consolidate_account(_petty_cash_account(), tmp_path)

    karibu_path = tmp_path / "Statements" / "Petty Cash UGX" / "Petty Cash UGX Karibu Ledger - 2026.xlsx"
    wb = load_workbook(karibu_path, read_only=True)
    sheet_names = list(wb.sheetnames)
    wb.close()

    # Karibu covers Jan..May → exactly those 5 month sheets.
    assert sheet_names == ["Jan", "Feb", "Mar", "Apr", "May"]


def test_consolidator_is_byte_identical_on_rerun(
    tmp_path: Path, momo_agent_xlsx, karibu_petty_cash_csv
):
    _stage_petty_cash_inputs(tmp_path, momo_agent_xlsx, karibu_petty_cash_csv)

    r1 = consolidate_account(_petty_cash_account(), tmp_path)
    h1 = {p.name: hashlib.sha256(p.read_bytes()).hexdigest()
          for p in r1.statement_workbooks_written + r1.karibu_workbooks_written}

    r2 = consolidate_account(_petty_cash_account(), tmp_path)
    h2 = {p.name: hashlib.sha256(p.read_bytes()).hexdigest()
          for p in r2.statement_workbooks_written + r2.karibu_workbooks_written}

    assert h1 == h2, "Two consolidator runs over identical inputs must produce byte-identical output"


def test_unparseable_date_row_surfaces_in_separate_sheet(tmp_path: Path):
    """A row with an unparseable date must appear in an Unparseable sheet,
    not silently dropped (Joash, 2026-05-20).
    """
    # Build a synthetic MoMo xlsx with one valid row and one bad-date row.
    tx_dir = tmp_path / "Transactions" / "Petty Cash UGX"
    tx_dir.mkdir(parents=True)
    momo = tx_dir / "synthetic.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    headers = ["Date / Time", "Transaction ID", "Transaction Type", "Amount",
               "From Account", "To Account", "Fee", "Commision Amount",
               "TAX", "Commision Receiving No.", "Commision Balance",
               "Float Balance"]
    ws.append(headers)
    ws.append(["2026-05-15 13:47", "111111111", "CASH_IN", 10000,
               "256770000001", "256770000002", 0, 0, 0, 0, 0, 0])
    # The bad row: nonsense date.
    ws.append(["NOT A DATE", "222222222", "CASH_IN", 20000,
               "256770000003", "256770000004", 0, 0, 0, 0, 0, 0])
    wb.save(momo)

    # Empty Karibu folder is fine — Karibu side is independent.
    (tmp_path / "Reports" / "Karibu" / "Petty Cash UGX").mkdir(parents=True)

    result = consolidate_account(_petty_cash_account(), tmp_path)
    assert result.statement_records_unique == 2
    assert result.statement_unparseable == 1

    out = tmp_path / "Statements" / "Petty Cash UGX" / "Petty Cash UGX Transactions - 2026.xlsx"
    wb = load_workbook(out, read_only=True)
    sheets = list(wb.sheetnames)
    wb.close()
    assert "Unparseable" in sheets, f"missing Unparseable sheet; got {sheets}"

    # The Unparseable row should carry the AUDIT_UNPARSEABLE_DATE flag.
    wb = load_workbook(out, read_only=True)
    ws = wb["Unparseable"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 1
    # `Audit Flag` is the last column in the statement layout.
    assert rows[0][-1] == AUDIT_UNPARSEABLE_DATE
    wb.close()


def test_consolidator_recovers_from_stale_nat_baseline(tmp_path: Path):
    """The new consolidator reads only source CSVs — a stale legacy xlsx
    in Statements/ cannot poison its output. This is the structural fix
    for the April-6 NaT-baseline bug.
    """
    # Build a synthetic MoMo xlsx with a valid post-April-6 row.
    tx_dir = tmp_path / "Transactions" / "Petty Cash UGX"
    tx_dir.mkdir(parents=True)
    momo = tx_dir / "fresh.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Date / Time", "Transaction ID", "Transaction Type", "Amount",
               "From Account", "To Account", "Fee", "Commision Amount",
               "TAX", "Commision Receiving No.", "Commision Balance",
               "Float Balance"])
    ws.append(["2026-04-20 09:00", "999999", "CASH_IN", 50000,
               "256770000010", "256770000020", 0, 0, 0, 0, 0, 0])
    wb.save(momo)

    # Pre-seed Statements/ with a "broken" legacy file. The consolidator
    # must IGNORE this — it only reads from Transactions/.
    statements = tmp_path / "Statements" / "Petty Cash UGX"
    statements.mkdir(parents=True)
    poison = statements / "Petty Cash UGX Transactions - 2026.xlsx"
    poison.write_text("THIS IS NOT A VALID XLSX, MUST BE OVERWRITTEN")

    (tmp_path / "Reports" / "Karibu" / "Petty Cash UGX").mkdir(parents=True)

    result = consolidate_account(_petty_cash_account(), tmp_path)
    assert result.statement_records_unique == 1
    # The poison file should have been overwritten with a valid xlsx.
    wb = load_workbook(poison, read_only=True)
    apr_sheet = wb["Apr"]
    rows = list(apr_sheet.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 1
    # 1st col is Date; 2nd is Transaction ID; check ID survived.
    assert str(rows[0][1]) == "999999"
    wb.close()


# ---------- dedupe helpers ----------

def _stmt_record(date, txn_id, amount, direction):
    from decimal import Decimal
    from parsers.types import NormalizedRecord
    return NormalizedRecord(
        source_file="x.csv", date=date, txn_id=txn_id,
        amount=Decimal(amount), direction=direction,
        counterparty="", txn_type="",
    )


def test_dedupe_collapses_identical_statement_rows():
    a = _stmt_record(datetime(2026, 4, 7), "A", 1000, "IN")
    b = _stmt_record(datetime(2026, 4, 7), "A", 1000, "IN")  # exact dup
    c = _stmt_record(datetime(2026, 4, 7), "B", 1000, "IN")
    result = _dedupe([a, b, c], _statement_dedup_key)
    assert len(result) == 2


def test_dedupe_keeps_all_nat_rows_separate():
    """Unparseable-date rows must not collapse against each other — the
    user needs to see every offending row in the review sheet."""
    a = _stmt_record(None, "A", 1000, "IN")
    b = _stmt_record(None, "A", 1000, "IN")
    result = _dedupe([a, b], _statement_dedup_key)
    assert len(result) == 2
