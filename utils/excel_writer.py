"""Excel output writer with BSR styling for reconciliation workbooks and statements."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
import pandas as pd

# ---------------------------------------------------------------------------
# Color constants
# ---------------------------------------------------------------------------
COLORS = {
    "mtn_header_bg": "1F6B2E",
    "airtel_header_bg": "C0392B",
    "header_text": "FFFFFF",
    "matched_bg": "D6EFDD",
    "not_in_statement_bg": "FDECEA",
    "not_in_karibu_bg": "FFF3CD",
    "flagged_bg": "FCE4EC",
    "zebra_mtn": "F0F7F0",
    "zebra_airtel": "FDF2F0",
    "audit_flag_bg": "FF9800",
    "status_matched": "1A6B2E",
    "status_not_in_stmt": "C0392B",
    "status_not_in_karibu": "B7791A",
    "conf_high": "1A6B2E",
    "conf_mid": "2471A3",
    "conf_low": "B7791A",
    "conf_very_low": "C0392B",
}

THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)


def _header_fill(channel: str) -> PatternFill:
    color = COLORS["mtn_header_bg"] if channel.upper() == "MTN" else COLORS["airtel_header_bg"]
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _zebra_fill(channel: str) -> PatternFill:
    color = COLORS["zebra_mtn"] if channel.upper() == "MTN" else COLORS["zebra_airtel"]
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _status_font(status: str) -> Font:
    color_map = {
        "Matched": COLORS["status_matched"],
        "Not in Statement": COLORS["status_not_in_stmt"],
        "Not in Karibu": COLORS["status_not_in_karibu"],
    }
    c = color_map.get(status, "000000")
    return Font(name="Arial", size=9, bold=True, color=c)


def _confidence_font(conf_str: str) -> Font:
    if not conf_str or conf_str == "—":
        return Font(name="Arial", size=9)
    try:
        val = int(conf_str.replace("%", ""))
    except ValueError:
        return Font(name="Arial", size=9)
    if val >= 90:
        c = COLORS["conf_high"]
    elif val >= 70:
        c = COLORS["conf_mid"]
    elif val >= 50:
        c = COLORS["conf_low"]
    else:
        c = COLORS["conf_very_low"]
    return Font(name="Arial", size=9, bold=True, color=c)


def _row_fill(status: str, has_flag: bool) -> PatternFill | None:
    if has_flag:
        return PatternFill(start_color=COLORS["flagged_bg"], end_color=COLORS["flagged_bg"], fill_type="solid")
    fill_map = {
        "Matched": PatternFill(start_color=COLORS["matched_bg"], end_color=COLORS["matched_bg"], fill_type="solid"),
        "Not in Statement": PatternFill(start_color=COLORS["not_in_statement_bg"], end_color=COLORS["not_in_statement_bg"], fill_type="solid"),
        "Not in Karibu": PatternFill(start_color=COLORS["not_in_karibu_bg"], end_color=COLORS["not_in_karibu_bg"], fill_type="solid"),
    }
    return fill_map.get(status)


# ---------------------------------------------------------------------------
# Statement writer
# ---------------------------------------------------------------------------

def write_mtn_statement(df: pd.DataFrame, file_path, banner_text: str):
    """Write MTN consolidated statement with BSR styling."""
    _write_statement(df, file_path, banner_text, "MTN Transactions", "MTN")


def write_airtel_statement(df: pd.DataFrame, file_path, banner_text: str):
    """Write Airtel consolidated statement with BSR styling."""
    _write_statement(df, file_path, banner_text, "All Transactions", "Airtel")


def _write_statement(df: pd.DataFrame, file_path, banner_text: str, sheet_name: str, channel: str):
    """Generic statement writer."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    headers = list(df.columns)
    ncols = len(headers)

    # Row 1: Banner
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    cell = ws.cell(row=1, column=1, value=banner_text)
    cell.font = Font(name="Arial", size=10, bold=True, italic=True)
    cell.alignment = Alignment(horizontal="left", vertical="center")

    # Row 2: Headers
    hfill = _header_fill(channel)
    hfont = Font(name="Arial", size=10, bold=True, color=COLORS["header_text"])
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    # Data rows
    data_font = Font(name="Arial", size=9)
    zfill = _zebra_fill(channel)
    for row_idx, (_, row) in enumerate(df.iterrows(), 3):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = THIN_BORDER
            if (row_idx - 3) % 2 == 1:
                cell.fill = zfill

    # Amount columns — number format
    amount_cols = _find_cols(headers, ["Amount", "Transaction Amount", "Balance",
                                        "Total Service Charge", "DR", "CR"])
    for col_idx in amount_cols:
        for row_idx in range(3, ws.max_row + 1):
            ws.cell(row=row_idx, column=col_idx).number_format = "#,##0"

    # Date columns — date format
    date_cols = _find_cols(headers, ["Date", "Transaction Date"])
    for col_idx in date_cols:
        for row_idx in range(3, ws.max_row + 1):
            ws.cell(row=row_idx, column=col_idx).number_format = "DD/MM/YYYY"

    # Freeze panes: row 3 (banner + header frozen)
    ws.freeze_panes = "A3"

    # Auto-width
    _auto_width(ws)

    wb.save(file_path)


# ---------------------------------------------------------------------------
# Reconciliation writer
# ---------------------------------------------------------------------------

def write_reconciliation(
    karibu_df: pd.DataFrame,
    stmt_df: pd.DataFrame,
    dashboard_lines: list[str],
    file_path,
    channel: str,
):
    """Write reconciliation workbook with Karibu Report, Merchant Statement, and Dashboard sheets."""
    wb = Workbook()

    # Sheet 1: Karibu Report
    ws1 = wb.active
    ws1.title = "Karibu Report"
    _write_recon_sheet(ws1, karibu_df, channel)

    # Sheet 2: Merchant Statement
    ws2 = wb.create_sheet("Merchant Statement")
    _write_recon_sheet(ws2, stmt_df, channel)

    # Sheet 3: Dashboard
    ws3 = wb.create_sheet("Dashboard")
    _write_dashboard(ws3, dashboard_lines, channel)

    wb.save(file_path)


def _write_recon_sheet(ws, df: pd.DataFrame, channel: str):
    """Write a reconciliation sheet (Karibu or Statement) with styling."""
    if df.empty:
        ws.cell(row=1, column=1, value="No data")
        return

    headers = list(df.columns)
    ncols = len(headers)

    # Find key column indices
    status_col = _find_col_idx(headers, "Status")
    conf_col = _find_col_idx(headers, "Confidence")
    flag_col = _find_col_idx(headers, "Audit Flag")

    # Row 1: Headers
    hfill = _header_fill(channel)
    hfont = Font(name="Arial", size=10, bold=True, color=COLORS["header_text"])
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    # Data rows
    data_font = Font(name="Arial", size=9)
    zfill = _zebra_fill(channel)
    flag_fill = PatternFill(start_color=COLORS["audit_flag_bg"], end_color=COLORS["audit_flag_bg"], fill_type="solid")
    flag_font = Font(name="Arial", size=9, bold=True, color="000000")

    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        status_val = str(row.get("Status", "")) if status_col else ""
        has_flag = bool(row.get("Audit Flag", "")) and str(row.get("Audit Flag", "")) not in ("", "nan", "—")
        row_bg = _row_fill(status_val, has_flag)

        for col_idx, (col_name, val) in enumerate(zip(headers, row), 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val if not pd.isna(val) else "")
            cell.border = THIN_BORDER

            # Apply row background or zebra
            if row_bg:
                cell.fill = row_bg
            elif (row_idx - 2) % 2 == 1:
                cell.fill = zfill

            # Status column styling
            if col_name == "Status":
                cell.font = _status_font(str(val) if not pd.isna(val) else "")
            elif col_name == "Confidence":
                cell.font = _confidence_font(str(val) if not pd.isna(val) else "")
            elif col_name == "Audit Flag" and has_flag:
                cell.fill = flag_fill
                cell.font = flag_font
            else:
                cell.font = data_font

    # Number format for amount columns
    amount_cols = _find_cols(headers, ["DR (UGX)", "CR (UGX)", "Balance", "Amount (UGX)",
                                        "Transaction Amount", "DR", "CR"])
    for col_idx in amount_cols:
        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row=row_idx, column=col_idx).number_format = "#,##0"

    # Date format
    date_cols = _find_cols(headers, ["Date", "Transaction Date"])
    for col_idx in date_cols:
        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row=row_idx, column=col_idx).number_format = "DD/MM/YYYY"

    ws.freeze_panes = "A2"
    _auto_width(ws)


def _write_dashboard(ws, lines: list[str], channel: str):
    """Write dashboard text content."""
    title_font = Font(name="Arial", size=14, bold=True,
                      color=COLORS["mtn_header_bg"] if channel.upper() == "MTN" else COLORS["airtel_header_bg"])
    section_font = Font(name="Arial", size=11, bold=True)
    data_font = Font(name="Arial", size=10)

    for i, line in enumerate(lines, 1):
        cell = ws.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = title_font
        elif line and not line.startswith(" "):
            cell.font = section_font
        else:
            cell.font = data_font

    ws.column_dimensions["A"].width = 60


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _find_col_idx(headers: list, name: str) -> int | None:
    """Find 1-based column index by header name."""
    for i, h in enumerate(headers):
        if h == name:
            return i + 1
    return None


def _find_cols(headers: list, names: list[str]) -> list[int]:
    """Find all 1-based column indices matching any of the given names."""
    result = []
    for i, h in enumerate(headers):
        if h in names:
            result.append(i + 1)
    return result


def _auto_width(ws, max_width: int = 30):
    """Auto-adjust column widths based on content."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col[:50]:  # Sample first 50 rows
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)
