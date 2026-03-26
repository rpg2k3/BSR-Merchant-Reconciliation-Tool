"""Auto-detection logic for uploaded file types.

Identifies MTN/Airtel transaction CSVs, Karibu ledger CSVs, and statement Excel files.
"""

from pathlib import Path

# Possible detected types and their destination folders (relative to WORKING_DIR)
FILE_TYPES = {
    "MTN Transaction":       "Transactions/MTN",
    "Airtel Transaction":    "Transactions/Airtel",
    "Karibu MTN Report":     "Reports/Karibu/MTN",
    "Karibu Airtel Report":  "Reports/Karibu/Airtel",
    "MTN Statement":         "Statements",
    "Airtel Statement":      "Statements",
}


def detect_file_type(file_path: Path) -> str | None:
    """Detect the type of a CSV or Excel file.

    Returns one of the FILE_TYPES keys, or None if unrecognised.
    """
    name = file_path.name.lower()
    suffix = file_path.suffix.lower()

    if suffix == ".xlsx":
        return _detect_xlsx(file_path)
    elif suffix == ".csv":
        return _detect_csv(file_path, name)
    return None


def _detect_xlsx(file_path: Path) -> str | None:
    """Detect statement Excel files by sheet name."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True)
        sheets = wb.sheetnames
        wb.close()
        if "MTN Transactions" in sheets:
            return "MTN Statement"
        if "All Transactions" in sheets:
            return "Airtel Statement"
    except Exception:
        pass
    return None


def _detect_csv(file_path: Path, name: str) -> str | None:
    """Detect CSV file type by reading header lines."""
    try:
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            head_lines = [f.readline() for _ in range(8)]
    except OSError:
        return None

    head_text = "".join(head_lines).lower()

    # Airtel Customer Transaction Report
    if "customer_transaction_report" in head_text or "customer transaction report" in head_text:
        return "Airtel Transaction"

    # Airtel User Transaction Report
    if "user_transaction_report" in head_text or "user transaction report" in head_text:
        return "Airtel Transaction"

    # Karibu HMS Ledger — check for "Ledger statement" header
    if "ledger statement" in head_text:
        # Determine MTN vs Airtel by reading account column values
        return _detect_karibu_channel(file_path)

    # MTN Transaction CSV — first line is header with Id,External id,Date,...
    first_line = head_lines[0].strip().lower() if head_lines else ""
    if first_line.startswith("id,") and "date," in first_line and "amount," in first_line:
        return "MTN Transaction"

    # Fallback: filename heuristics
    if "customer" in name and "transaction" in name:
        return "Airtel Transaction"
    if "user" in name and "transaction" in name:
        return "Airtel Transaction"
    if "karibu" in name or "ledger" in name:
        if "mtn" in name:
            return "Karibu MTN Report"
        if "airtel" in name:
            return "Karibu Airtel Report"
    # MTN merchant ID pattern in filename (e.g. 187819368_)
    import re
    if re.search(r"\d{9}_", name):
        return "MTN Transaction"
    if "mtn" in name:
        if "recon" not in name and "statement" not in name:
            return "Karibu MTN Report"
    if "airtel" in name:
        if "recon" not in name and "statement" not in name:
            return "Karibu Airtel Report"

    return None


def _detect_karibu_channel(file_path: Path) -> str:
    """Read a Karibu CSV to determine if it's MTN or Airtel by Account column."""
    try:
        import csv
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            # Skip first 2 rows (title + blank)
            f.readline()
            f.readline()
            reader = csv.DictReader(f)
            for row in reader:
                account = (row.get("Account", "") or "").strip().lower()
                if not account or account == "":
                    continue
                if "mtn" in account:
                    return "Karibu MTN Report"
                if "airtel" in account:
                    return "Karibu Airtel Report"
    except Exception:
        pass
    # Fallback: check filename
    name = file_path.name.lower()
    if "mtn" in name:
        return "Karibu MTN Report"
    if "airtel" in name:
        return "Karibu Airtel Report"
    return "Karibu MTN Report"  # default
