"""Module 2 — Reconciliation Engine.

Multi-pass matching between merchant statements and Karibu HMS ledger reports.
"""

from pathlib import Path
from datetime import datetime, timedelta
from itertools import combinations

import numpy as np
import pandas as pd

from core.parsers import (
    load_mtn_statement,
    load_airtel_statement,
    load_all_karibu,
)
from core.anomalies import run_anomaly_detection
from utils.backup import create_backup
from utils.excel_writer import write_reconciliation


def reconcile(channel: str, base_dir: Path, config: dict, log_fn=None) -> dict:
    """Run reconciliation for MTN or Airtel.

    Returns dict with summary statistics.
    """
    def log(msg, level="info"):
        if log_fn:
            log_fn(msg, level)

    recon_path = base_dir / "Reconciliation" / f"BSR_{channel}_Reconciliation.xlsx"
    backup_dir = base_dir / "Backups"
    date_tolerance = config.get("date_tolerance_days", 2)

    # Backup existing reconciliation
    backup_path = create_backup(recon_path, backup_dir)
    if backup_path:
        log(f"Backup saved: {backup_path}")

    # Load existing comments to preserve
    existing_comments = _load_existing_comments(recon_path)

    # Load statement
    stmt_path = base_dir / "Statements" / f"BSR_{channel}_Merchant_Transactions.xlsx"
    if not stmt_path.exists():
        log(f"Statement file not found: {stmt_path}", "error")
        return {"error": "Statement file not found"}

    if channel == "MTN":
        stmt_df, _ = load_mtn_statement(stmt_path)
        stmt_id_col = "Id"
        stmt_date_col = "Date"
        stmt_amount_col = "Amount"
        stmt_payer_col = "From name"
    else:
        stmt_df, _ = load_airtel_statement(stmt_path)
        stmt_id_col = "Transaction ID"
        stmt_date_col = "Transaction Date"
        stmt_amount_col = "Transaction Amount"
        stmt_payer_col = "Payer User Name"

    log(f"Loaded {channel} statement: {len(stmt_df)} rows")

    # Load Karibu
    karibu_dir = base_dir / "Reports" / "Karibu" / channel
    karibu_df = load_all_karibu(karibu_dir)
    if karibu_df.empty:
        log(f"No Karibu data found in {karibu_dir}", "error")
        return {"error": "No Karibu data"}

    log(f"Loaded Karibu ledger: {len(karibu_df)} rows")

    # Filter statement: positive amounts only for matching (keep contras for output)
    stmt_df[stmt_amount_col] = pd.to_numeric(stmt_df[stmt_amount_col], errors="coerce")
    stmt_df[stmt_date_col] = pd.to_datetime(stmt_df[stmt_date_col], errors="coerce")

    # Prepare Karibu: filter DR > 0 for matching
    karibu_dr = karibu_df[karibu_df["DR"] > 0].copy()
    karibu_dr.reset_index(drop=True, inplace=True)
    log(f"Karibu DR entries for matching: {len(karibu_dr)}")

    # Statement positive amounts for matching
    stmt_positive = stmt_df[stmt_df[stmt_amount_col] > 0].copy()
    stmt_positive.reset_index(drop=True, inplace=True)
    log(f"Statement positive entries for matching: {len(stmt_positive)}")

    # Initialize match tracking arrays
    karibu_matched = [False] * len(karibu_dr)
    stmt_matched = [False] * len(stmt_positive)

    # Result columns for Karibu
    karibu_status = [""] * len(karibu_dr)
    karibu_match_type = [""] * len(karibu_dr)
    karibu_confidence = [""] * len(karibu_dr)
    karibu_matched_ref = [""] * len(karibu_dr)

    # Result columns for Statement
    stmt_status = [""] * len(stmt_positive)
    stmt_match_type = [""] * len(stmt_positive)
    stmt_confidence = [""] * len(stmt_positive)
    stmt_matched_ref = [""] * len(stmt_positive)

    # -------------------------------------------------------------------
    # Pass 1-3: Exact amount matches with varying date tolerance
    # -------------------------------------------------------------------
    passes = [
        (0, "100%", "Exact"),   # Same date
        (1, "90%", "Exact"),    # ±1 day
        (2, "80%", "Exact"),    # ±2 days
    ]
    # Extend up to user-configured tolerance
    if date_tolerance > 2:
        for d in range(3, date_tolerance + 1):
            conf = max(70, 80 - (d - 2) * 10)
            passes.append((d, f"{conf}%", "Exact"))

    for max_days, conf, mtype in passes:
        for ki in range(len(karibu_dr)):
            if karibu_matched[ki]:
                continue
            k_amount = karibu_dr.iloc[ki]["DR"]
            k_date = karibu_dr.iloc[ki]["Date"]
            if pd.isna(k_date) or pd.isna(k_amount):
                continue

            for si in range(len(stmt_positive)):
                if stmt_matched[si]:
                    continue
                s_amount = stmt_positive.iloc[si][stmt_amount_col]
                s_date = stmt_positive.iloc[si][stmt_date_col]
                if pd.isna(s_date) or pd.isna(s_amount):
                    continue

                if abs(float(k_amount) - float(s_amount)) < 0.5:
                    day_diff = abs((pd.Timestamp(k_date) - pd.Timestamp(s_date)).days)
                    if day_diff <= max_days:
                        # Match found
                        karibu_matched[ki] = True
                        stmt_matched[si] = True
                        s_id = str(stmt_positive.iloc[si].get(stmt_id_col, si))
                        karibu_status[ki] = "Matched"
                        karibu_match_type[ki] = mtype
                        karibu_confidence[ki] = conf
                        karibu_matched_ref[ki] = s_id
                        stmt_status[si] = "Matched"
                        stmt_match_type[si] = mtype
                        stmt_confidence[si] = conf
                        stmt_matched_ref[si] = f"K{ki}"
                        break

    matched_exact = sum(karibu_matched)
    log(f"After exact passes: {matched_exact} Karibu rows matched")

    # -------------------------------------------------------------------
    # Pass 4-5: Lumpsum K→S (one Karibu = sum of multiple stmt rows)
    # -------------------------------------------------------------------
    for max_days, conf in [(0, "60%"), (date_tolerance, "45%")]:
        for ki in range(len(karibu_dr)):
            if karibu_matched[ki]:
                continue
            k_amount = karibu_dr.iloc[ki]["DR"]
            k_date = karibu_dr.iloc[ki]["Date"]
            if pd.isna(k_date) or pd.isna(k_amount) or k_amount <= 0:
                continue

            # Find unmatched stmt rows within date window
            candidates = []
            for si in range(len(stmt_positive)):
                if stmt_matched[si]:
                    continue
                s_date = stmt_positive.iloc[si][stmt_date_col]
                s_amount = stmt_positive.iloc[si][stmt_amount_col]
                if pd.isna(s_date) or pd.isna(s_amount) or s_amount <= 0:
                    continue
                day_diff = abs((pd.Timestamp(k_date) - pd.Timestamp(s_date)).days)
                if day_diff <= max_days:
                    candidates.append((si, float(s_amount)))

            if not candidates:
                continue

            # Greedy subset-sum: find combination that sums to k_amount
            target = float(k_amount)
            subset = _greedy_subset_sum(candidates, target)
            if subset:
                karibu_matched[ki] = True
                refs = []
                for si, _ in subset:
                    stmt_matched[si] = True
                    s_id = str(stmt_positive.iloc[si].get(stmt_id_col, si))
                    refs.append(s_id)
                    stmt_status[si] = "Matched"
                    stmt_match_type[si] = "Lumpsum"
                    stmt_confidence[si] = conf
                    stmt_matched_ref[si] = f"K{ki}"

                karibu_status[ki] = "Matched"
                karibu_match_type[ki] = "Lumpsum"
                karibu_confidence[ki] = conf
                karibu_matched_ref[ki] = ",".join(refs)

    matched_lumpsum_ks = sum(karibu_matched) - matched_exact
    log(f"After lumpsum K→S: {matched_lumpsum_ks} additional Karibu rows matched")

    # -------------------------------------------------------------------
    # Pass 6: Lumpsum S→K (one stmt = sum of multiple Karibu DRs)
    # -------------------------------------------------------------------
    for si in range(len(stmt_positive)):
        if stmt_matched[si]:
            continue
        s_amount = stmt_positive.iloc[si][stmt_amount_col]
        s_date = stmt_positive.iloc[si][stmt_date_col]
        if pd.isna(s_date) or pd.isna(s_amount) or s_amount <= 0:
            continue

        candidates = []
        for ki in range(len(karibu_dr)):
            if karibu_matched[ki]:
                continue
            k_date = karibu_dr.iloc[ki]["Date"]
            k_amount = karibu_dr.iloc[ki]["DR"]
            if pd.isna(k_date) or pd.isna(k_amount) or k_amount <= 0:
                continue
            day_diff = abs((pd.Timestamp(s_date) - pd.Timestamp(k_date)).days)
            if day_diff <= 0:
                candidates.append((ki, float(k_amount)))

        if not candidates:
            continue

        target = float(s_amount)
        subset = _greedy_subset_sum(candidates, target)
        if subset:
            stmt_matched[si] = True
            refs = []
            s_id = str(stmt_positive.iloc[si].get(stmt_id_col, si))
            for ki, _ in subset:
                karibu_matched[ki] = True
                refs.append(f"K{ki}")
                karibu_status[ki] = "Matched"
                karibu_match_type[ki] = "Lumpsum"
                karibu_confidence[ki] = "55%"
                karibu_matched_ref[ki] = s_id

            stmt_status[si] = "Matched"
            stmt_match_type[si] = "Lumpsum"
            stmt_confidence[si] = "55%"
            stmt_matched_ref[si] = ",".join(refs)

    # -------------------------------------------------------------------
    # Pass 7: Amount only (any date difference)
    # -------------------------------------------------------------------
    for ki in range(len(karibu_dr)):
        if karibu_matched[ki]:
            continue
        k_amount = karibu_dr.iloc[ki]["DR"]
        if pd.isna(k_amount):
            continue

        for si in range(len(stmt_positive)):
            if stmt_matched[si]:
                continue
            s_amount = stmt_positive.iloc[si][stmt_amount_col]
            if pd.isna(s_amount):
                continue

            if abs(float(k_amount) - float(s_amount)) < 0.5:
                karibu_matched[ki] = True
                stmt_matched[si] = True
                s_id = str(stmt_positive.iloc[si].get(stmt_id_col, si))
                karibu_status[ki] = "Matched"
                karibu_match_type[ki] = "Amount Only"
                karibu_confidence[ki] = "40%"
                karibu_matched_ref[ki] = s_id
                stmt_status[si] = "Matched"
                stmt_match_type[si] = "Amount Only"
                stmt_confidence[si] = "40%"
                stmt_matched_ref[si] = f"K{ki}"
                break

    # -------------------------------------------------------------------
    # Label unmatched
    # -------------------------------------------------------------------
    for ki in range(len(karibu_dr)):
        if not karibu_matched[ki]:
            karibu_status[ki] = "Not in Statement"
            karibu_match_type[ki] = "—"
            karibu_confidence[ki] = "—"
            karibu_matched_ref[ki] = "—"

    for si in range(len(stmt_positive)):
        if not stmt_matched[si]:
            stmt_status[si] = "Not in Karibu"
            stmt_match_type[si] = "—"
            stmt_confidence[si] = "—"
            stmt_matched_ref[si] = "—"

    # -------------------------------------------------------------------
    # Build output DataFrames
    # -------------------------------------------------------------------

    # Karibu output sheet
    karibu_out = pd.DataFrame({
        "Date": karibu_dr["Date"],
        "Account": karibu_dr["Account"],
        "Narration": karibu_dr["Narration"],
        "DR (UGX)": karibu_dr["DR"],
        "CR (UGX)": karibu_dr.get("CR", 0),
        "Balance": karibu_dr.get("Balance", ""),
        "Status": karibu_status,
        "Match Type": karibu_match_type,
        "Confidence": karibu_confidence,
        "Matched Ref": karibu_matched_ref,
        "Audit Flag": "",
        "Comments": "",
    })

    # Statement output sheet
    if channel == "MTN":
        stmt_out = pd.DataFrame({
            "Date": stmt_positive[stmt_date_col],
            "Transaction ID": stmt_positive[stmt_id_col],
            "Payer Name": stmt_positive.get(stmt_payer_col, ""),
            "Amount (UGX)": stmt_positive[stmt_amount_col],
            "Tx Status": stmt_positive.get("Status", ""),
            "Status": stmt_status,
            "Match Type": stmt_match_type,
            "Confidence": stmt_confidence,
            "Matched Ref": stmt_matched_ref,
            "Audit Flag": "",
            "Comments": "",
        })
    else:
        stmt_out = pd.DataFrame({
            "Date": stmt_positive[stmt_date_col],
            "Transaction ID": stmt_positive[stmt_id_col],
            "Payer Name": stmt_positive.get(stmt_payer_col, ""),
            "Amount (UGX)": stmt_positive[stmt_amount_col],
            "Reference": stmt_positive.get("Reference No.", ""),
            "Tx Status": stmt_positive.get("Transaction Status", ""),
            "Status": stmt_status,
            "Match Type": stmt_match_type,
            "Confidence": stmt_confidence,
            "Matched Ref": stmt_matched_ref,
            "Audit Flag": "",
            "Comments": "",
        })

    # Also add contra/negative rows to statement output (unmatched, flagged)
    stmt_contras = stmt_df[stmt_df[stmt_amount_col] < 0].copy() if channel == "MTN" else \
        stmt_df[stmt_df.get("Service Type", pd.Series(dtype=str)).str.contains("Contra", case=False, na=False)].copy()
    if not stmt_contras.empty:
        if channel == "MTN":
            contra_out = pd.DataFrame({
                "Date": stmt_contras[stmt_date_col],
                "Transaction ID": stmt_contras[stmt_id_col],
                "Payer Name": stmt_contras.get(stmt_payer_col, ""),
                "Amount (UGX)": stmt_contras[stmt_amount_col],
                "Tx Status": stmt_contras.get("Status", ""),
                "Status": "Contra",
                "Match Type": "—",
                "Confidence": "—",
                "Matched Ref": "—",
                "Audit Flag": "",
                "Comments": "",
            })
        else:
            contra_out = pd.DataFrame({
                "Date": stmt_contras[stmt_date_col],
                "Transaction ID": stmt_contras[stmt_id_col],
                "Payer Name": stmt_contras.get(stmt_payer_col, ""),
                "Amount (UGX)": stmt_contras[stmt_amount_col],
                "Reference": stmt_contras.get("Reference No.", ""),
                "Tx Status": stmt_contras.get("Transaction Status", ""),
                "Status": "Contra",
                "Match Type": "—",
                "Confidence": "—",
                "Matched Ref": "—",
                "Audit Flag": "",
                "Comments": "",
            })
        stmt_out = pd.concat([stmt_out, contra_out], ignore_index=True)

    # Run anomaly detection
    log("Running anomaly detection...")
    karibu_out, stmt_out = run_anomaly_detection(
        karibu_out, stmt_out, karibu_df, stmt_df,
        channel, config, stmt_amount_col, stmt_date_col, stmt_payer_col
    )

    # Restore preserved comments
    _restore_comments(karibu_out, stmt_out, existing_comments, channel)

    # Build dashboard
    dashboard = _build_dashboard(karibu_out, stmt_out, karibu_df, stmt_df,
                                  channel, stmt_amount_col, stmt_date_col)

    # Write output
    log("Writing reconciliation workbook...")
    write_reconciliation(karibu_out, stmt_out, dashboard, recon_path, channel)
    log(f"Reconciliation saved: {recon_path}")

    # Summary stats
    total_karibu = len(karibu_dr)
    matched_karibu = sum(1 for s in karibu_status if s == "Matched")
    not_in_stmt = sum(1 for s in karibu_status if s == "Not in Statement")
    not_in_karibu = sum(1 for s in stmt_status if s == "Not in Karibu")

    log(f"Results: {matched_karibu}/{total_karibu} Karibu rows matched "
        f"({matched_karibu/total_karibu*100:.1f}%)" if total_karibu > 0 else "No Karibu data")
    log(f"  Not in Statement: {not_in_stmt}")
    log(f"  Not in Karibu: {not_in_karibu}")

    return {
        "matched": matched_karibu,
        "total_karibu": total_karibu,
        "not_in_statement": not_in_stmt,
        "not_in_karibu": not_in_karibu,
        "recon_path": str(recon_path),
    }


# ---------------------------------------------------------------------------
# Subset-sum helper
# ---------------------------------------------------------------------------

def _greedy_subset_sum(candidates: list[tuple[int, float]], target: float, tolerance: float = 0.5) -> list | None:
    """Find a subset of candidates that sums to target within tolerance.

    candidates: list of (index, amount) tuples.
    Uses greedy approach: try largest first, then brute-force small combos.
    """
    if not candidates:
        return None

    # Sort by amount descending
    sorted_cands = sorted(candidates, key=lambda x: x[1], reverse=True)

    # Quick check: any single candidate matches?
    for c in sorted_cands:
        if abs(c[1] - target) < tolerance:
            return [c]

    # Greedy: pick largest that fits, repeat
    remaining = target
    selected = []
    unused = list(sorted_cands)

    for c in sorted_cands:
        if c[1] <= remaining + tolerance:
            selected.append(c)
            remaining -= c[1]
            if abs(remaining) < tolerance:
                return selected

    # If greedy failed, try small combinations (up to 5 items, limited to 15 candidates)
    trimmed = sorted_cands[:15]
    for r in range(2, min(6, len(trimmed) + 1)):
        for combo in combinations(trimmed, r):
            total = sum(c[1] for c in combo)
            if abs(total - target) < tolerance:
                return list(combo)

    return None


# ---------------------------------------------------------------------------
# Comment preservation
# ---------------------------------------------------------------------------

def _load_existing_comments(recon_path: Path) -> dict:
    """Load existing Comments from a reconciliation file for preservation."""
    comments = {"karibu": {}, "stmt": {}}
    if not recon_path.exists():
        return comments

    try:
        from openpyxl import load_workbook
        wb = load_workbook(recon_path, data_only=True)

        if "Karibu Report" in wb.sheetnames:
            ws = wb["Karibu Report"]
            headers = [cell.value for cell in ws[1]]
            comment_idx = headers.index("Comments") if "Comments" in headers else None
            date_idx = headers.index("Date") if "Date" in headers else None
            narr_idx = headers.index("Narration") if "Narration" in headers else None
            dr_idx = None
            for i, h in enumerate(headers):
                if h and "DR" in str(h):
                    dr_idx = i
                    break

            if comment_idx is not None:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    comment_val = row[comment_idx] if comment_idx < len(row) else None
                    if comment_val:
                        key_parts = []
                        if date_idx is not None and date_idx < len(row):
                            key_parts.append(str(row[date_idx]))
                        if narr_idx is not None and narr_idx < len(row):
                            key_parts.append(str(row[narr_idx]))
                        if dr_idx is not None and dr_idx < len(row):
                            key_parts.append(str(row[dr_idx]))
                        if key_parts:
                            comments["karibu"]["|".join(key_parts)] = comment_val

        if "Merchant Statement" in wb.sheetnames:
            ws = wb["Merchant Statement"]
            headers = [cell.value for cell in ws[1]]
            comment_idx = headers.index("Comments") if "Comments" in headers else None
            txid_idx = headers.index("Transaction ID") if "Transaction ID" in headers else None

            if comment_idx is not None and txid_idx is not None:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    comment_val = row[comment_idx] if comment_idx < len(row) else None
                    txid_val = row[txid_idx] if txid_idx < len(row) else None
                    if comment_val and txid_val:
                        comments["stmt"][str(txid_val)] = comment_val

        wb.close()
    except Exception:
        pass

    return comments


def _restore_comments(karibu_out: pd.DataFrame, stmt_out: pd.DataFrame,
                      comments: dict, channel: str):
    """Restore previously saved comments to matching rows."""
    if comments["karibu"]:
        for idx, row in karibu_out.iterrows():
            key = f"{row.get('Date', '')}|{row.get('Narration', '')}|{row.get('DR (UGX)', '')}"
            if key in comments["karibu"]:
                karibu_out.at[idx, "Comments"] = comments["karibu"][key]

    if comments["stmt"]:
        for idx, row in stmt_out.iterrows():
            txid = str(row.get("Transaction ID", ""))
            if txid in comments["stmt"]:
                stmt_out.at[idx, "Comments"] = comments["stmt"][txid]


# ---------------------------------------------------------------------------
# Dashboard builder
# ---------------------------------------------------------------------------

def _build_dashboard(karibu_out, stmt_out, karibu_full, stmt_full,
                     channel, stmt_amount_col, stmt_date_col) -> list[str]:
    """Build the dashboard summary text lines."""
    lines = []
    now = datetime.now().strftime("%d %b %Y")

    # Date range
    k_dates = karibu_out["Date"].dropna()
    s_dates = stmt_out["Date"].dropna()
    k_min = k_dates.min().strftime("%d %b %Y") if not k_dates.empty else "N/A"
    k_max = k_dates.max().strftime("%d %b %Y") if not k_dates.empty else "N/A"
    s_min = s_dates.min().strftime("%d %b %Y") if not s_dates.empty else "N/A"
    s_max = s_dates.max().strftime("%d %b %Y") if not s_dates.empty else "N/A"

    period_min = min(k_dates.min(), s_dates.min()).strftime("%d %b %Y") if not k_dates.empty and not s_dates.empty else "N/A"
    period_max = max(k_dates.max(), s_dates.max()).strftime("%d %b %Y") if not k_dates.empty and not s_dates.empty else "N/A"

    lines.append(f"BSR {channel} Reconciliation Dashboard")
    lines.append(f"Period: {period_min} – {period_max}     Generated: {now}")
    lines.append("")

    # Statement summary
    stmt_amounts = pd.to_numeric(stmt_full[stmt_amount_col], errors="coerce")
    total_received = stmt_amounts[stmt_amounts > 0].sum()
    total_contras = stmt_amounts[stmt_amounts < 0].sum()

    stmt_dates = pd.to_datetime(stmt_full[stmt_date_col], errors="coerce").dropna()

    lines.append("STATEMENT SUMMARY")
    lines.append(f"  Total transactions:        {len(stmt_full)}")
    lines.append(f"  Total received (UGX):      {total_received:,.0f}")
    lines.append(f"  Total contras (UGX):       {abs(total_contras):,.0f}")
    lines.append(f"  Net balance:               {total_received + total_contras:,.0f}")
    lines.append(f"  Date range:                {s_min} – {s_max}")
    lines.append("")

    # Karibu summary
    karibu_dr_total = pd.to_numeric(karibu_full["DR"], errors="coerce").fillna(0)
    karibu_cr_total = pd.to_numeric(karibu_full["CR"], errors="coerce").fillna(0)

    lines.append("KARIBU SUMMARY")
    lines.append(f"  Total ledger entries (DR): {len(karibu_full[karibu_full['DR'] > 0])}")
    lines.append(f"  Total DR value (UGX):      {karibu_dr_total.sum():,.0f}")
    lines.append(f"  Total CR/contra (UGX):     {karibu_cr_total.sum():,.0f}")
    lines.append(f"  Date range:                {k_min} – {k_max}")
    lines.append("")

    # Reconciliation results
    matched = len(karibu_out[karibu_out["Status"] == "Matched"])
    not_in_stmt = karibu_out[karibu_out["Status"] == "Not in Statement"]
    not_in_karibu = stmt_out[stmt_out["Status"] == "Not in Karibu"]
    total_k = len(karibu_out)
    pct = (matched / total_k * 100) if total_k > 0 else 0

    nis_value = pd.to_numeric(not_in_stmt["DR (UGX)"], errors="coerce").sum()
    nik_value = pd.to_numeric(not_in_karibu["Amount (UGX)"], errors="coerce").sum()
    variance = nis_value - nik_value

    lines.append("RECONCILIATION RESULTS")
    lines.append(f"  Matched:                   {matched} rows  ({pct:.1f}%)")
    lines.append(f"  Not in Statement:          {len(not_in_stmt)} rows  ({nis_value:,.0f} UGX)")
    lines.append(f"  Not in Karibu:             {len(not_in_karibu)} rows  ({nik_value:,.0f} UGX)")
    lines.append(f"  Unreconciled variance:     {variance:,.0f} UGX")
    lines.append("")

    # Match quality
    def _count_conf(df, low, high):
        count = 0
        for v in df["Confidence"]:
            try:
                val = int(str(v).replace("%", ""))
                if low <= val <= high:
                    count += 1
            except (ValueError, TypeError):
                pass
        return count

    lines.append("MATCH QUALITY")
    lines.append(f"  100% confidence:           {_count_conf(karibu_out, 100, 100)} rows")
    lines.append(f"  80-99% confidence:         {_count_conf(karibu_out, 80, 99)} rows")
    lines.append(f"  50-79% confidence:         {_count_conf(karibu_out, 50, 79)} rows")
    lines.append(f"  <50% confidence:           {_count_conf(karibu_out, 0, 49)} rows")
    lines.append("")

    # Audit flags summary
    all_flags = []
    for df in [karibu_out, stmt_out]:
        for v in df["Audit Flag"]:
            if v and str(v) not in ("", "nan"):
                all_flags.extend(str(v).split(","))
    flag_counts = {}
    for f in all_flags:
        f = f.strip()
        if f:
            flag_counts[f] = flag_counts.get(f, 0) + 1

    lines.append("AUDIT FLAGS SUMMARY")
    if flag_counts:
        for flag, count in sorted(flag_counts.items()):
            lines.append(f"  {flag}: {count} occurrences")
    else:
        lines.append("  No audit flags raised")
    lines.append("")

    # Contras section
    if channel == "MTN":
        contras_in_stmt = stmt_full[stmt_full["Amount"] < 0] if "Amount" in stmt_full.columns else pd.DataFrame()
    else:
        contras_in_stmt = stmt_out[stmt_out["Status"] == "Contra"]

    lines.append("CONTRAS")
    if not contras_in_stmt.empty and stmt_date_col in contras_in_stmt.columns:
        contra_dates = pd.to_datetime(contras_in_stmt[stmt_date_col], errors="coerce").dropna()
        if not contra_dates.empty:
            last_contra_date = contra_dates.max()
            last_row = contras_in_stmt.loc[contras_in_stmt[stmt_date_col] == last_contra_date]
            last_amount = pd.to_numeric(last_row.iloc[0][stmt_amount_col], errors="coerce") if not last_row.empty else 0
            days_since = (datetime.now() - last_contra_date).days

            lines.append(f"  Last contra date:          {last_contra_date.strftime('%d %b %Y')}")
            lines.append(f"  Last contra amount (UGX):  {abs(last_amount):,.0f}")
            lines.append(f"  Days since last contra:    {days_since}")
        else:
            lines.append("  No contra entries found")
    else:
        lines.append("  No contra entries found")

    return lines
