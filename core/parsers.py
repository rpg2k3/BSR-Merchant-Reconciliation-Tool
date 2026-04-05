"""Data parsers for BSR Reconciliation Tool.

Handles all CSV and Excel parsing for MTN, Airtel, and Karibu formats.
"""

import re
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# MTN Transaction CSV
# ---------------------------------------------------------------------------

def parse_mtn_csv(file_path: Path) -> pd.DataFrame:
    """Parse a raw MTN merchant portal CSV export.

    Standard CSV, no header rows to skip.
    Key columns: Id, Date, Amount, Status, From name, Type.
    Amount: negative = contra/withdrawal.
    """
    df = pd.read_csv(file_path, dtype={"Id": str})
    df["Id"] = df["Id"].astype(str).str.strip()
    df["Date"] = pd.to_datetime(df["Date"], format="%Y-%m-%d %H:%M:%S", errors="coerce")
    # Amount column — ensure numeric
    df["Amount"] = pd.to_numeric(df["Amount"], errors="coerce")
    return df


# ---------------------------------------------------------------------------
# Airtel Transaction CSVs
# ---------------------------------------------------------------------------

def _read_airtel_csv_flexible(file_path: Path, skiprows: int) -> pd.DataFrame:
    """Read an Airtel CSV, merging extra fields caused by unquoted commas.

    Airtel CSVs have free-text fields (e.g. Reference No.) that may contain
    commas without quoting, producing rows with more fields than columns.
    When extra fields are detected, they are merged back into the 3rd column
    (Reference No., 0-indexed col 2) which is the typical culprit.
    """
    import csv
    with open(file_path, "r", encoding="utf-8", errors="replace") as f:
        lines = f.readlines()
    # Skip header block
    lines = lines[skiprows:]
    if not lines:
        return pd.DataFrame()
    # Parse header to get expected column count
    header = next(csv.reader([lines[0]]))
    header = [h.strip() for h in header]
    n_cols = len(header)

    rows = []
    for line in lines[1:]:
        fields = next(csv.reader([line]))
        if len(fields) == n_cols:
            rows.append(fields)
        elif len(fields) > n_cols:
            # Merge the extra fields into the Reference No. column (index 2)
            extra = len(fields) - n_cols
            merged = ",".join(fields[2 : 2 + extra + 1])
            fixed = fields[:2] + [merged] + fields[2 + extra + 1 :]
            rows.append(fixed)
        elif len(fields) > 0:
            # Fewer fields than expected — pad with empty strings
            rows.append(fields + [""] * (n_cols - len(fields)))

    return pd.DataFrame(rows, columns=header, dtype=str)


def parse_airtel_customer_csv(file_path: Path) -> pd.DataFrame:
    """Parse Airtel Customer Transaction Report CSV.

    Skip first 5 rows (title block), row 6 is header (0-indexed: skiprows=5).
    Transaction ID may be in scientific notation — normalize to full integer string.
    """
    df = _read_airtel_csv_flexible(file_path, skiprows=5)
    # Clean column names
    df.columns = df.columns.str.strip()

    # Normalize Transaction ID from scientific notation
    if "Transaction ID" in df.columns:
        df["Transaction ID"] = df["Transaction ID"].apply(_normalize_airtel_id)

    # Parse date: DD-MMM-YYYY  HH:MM:SS (e.g. 21-MAR-2026  19:57:32)
    date_col = "Transaction Date & Time"
    if date_col in df.columns:
        df[date_col] = df[date_col].str.strip()
        df["Transaction Date"] = pd.to_datetime(
            df[date_col], format="%d-%b-%Y  %H:%M:%S", errors="coerce"
        )
        # Extract time component as string
        df["Transaction Time"] = df["Transaction Date"].dt.strftime("%H:%M:%S")
        # Keep date-only for the date column
        df["_parsed_datetime"] = df["Transaction Date"]
        df["Transaction Date"] = df["Transaction Date"].dt.normalize()
    else:
        df["Transaction Date"] = pd.NaT
        df["Transaction Time"] = ""
        df["_parsed_datetime"] = pd.NaT

    # Amount
    if "Transaction Amount" in df.columns:
        df["Transaction Amount"] = pd.to_numeric(
            df["Transaction Amount"].str.replace(",", ""), errors="coerce"
        )

    return df


def parse_airtel_user_csv(file_path: Path) -> pd.DataFrame:
    """Parse Airtel User Transaction Report CSV.

    Skip first 6 rows, row 7 is header (0-indexed: skiprows=6).
    Transaction Type: MP + ChannelWallet To Bank Transfer = contra.
    MR = merchant receipt. SCP = service charge (ignore).
    """
    df = _read_airtel_csv_flexible(file_path, skiprows=5)
    df.columns = df.columns.str.strip()

    if "Transaction ID" in df.columns:
        df["Transaction ID"] = df["Transaction ID"].apply(_normalize_airtel_id)

    # Parse date: DD-MMM-YY (e.g. 23-MAR-26)
    date_col = "Transaction Date and Time"
    if date_col in df.columns:
        df[date_col] = df[date_col].str.strip()
        df["Transaction Date"] = pd.to_datetime(
            df[date_col], format="%d-%b-%y", errors="coerce"
        )
    else:
        df["Transaction Date"] = pd.NaT

    if "Transaction Amount" in df.columns:
        df["Transaction Amount"] = pd.to_numeric(
            df["Transaction Amount"].str.replace(",", ""), errors="coerce"
        )

    # Filter: keep MP (contra) and MR (receipt), drop SCP (service charge)
    if "Transaction Type" in df.columns:
        df = df[df["Transaction Type"].isin(["MP", "MR"])].copy()

    return df


def _normalize_airtel_id(val) -> str:
    """Convert scientific notation IDs to full integer strings."""
    if pd.isna(val) or val is None:
        return ""
    val = str(val).strip()
    if "E" in val.upper() or "e" in val:
        try:
            return str(int(float(val)))
        except (ValueError, OverflowError):
            return val
    # Remove any decimal point for whole numbers
    try:
        f = float(val)
        if f == int(f):
            return str(int(f))
    except (ValueError, OverflowError):
        pass
    return val


def identify_airtel_csv_type(file_path: Path) -> str:
    """Identify whether an Airtel CSV is 'customer' or 'user' type.

    Reads the first few lines to check for identifying markers.
    """
    try:
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            head = f.read(500).lower()
        if "customer_transaction_report" in head or "customer transaction report" in head:
            return "customer"
        elif "user_transaction_report" in head or "user transaction report" in head:
            return "user"
    except OSError:
        pass
    # Fallback: guess from filename
    name = file_path.name.lower()
    if "customer" in name:
        return "customer"
    elif "user" in name:
        return "user"
    return "unknown"


# ---------------------------------------------------------------------------
# Consolidated Statement Excel (Statements folder)
# ---------------------------------------------------------------------------

def load_mtn_statement(file_path: Path) -> tuple[pd.DataFrame, str]:
    """Load MTN consolidated statement Excel.

    Sheet: 'MTN Transactions'
    Row 0: banner, Row 1: headers, Row 2+: data.
    Returns (dataframe, banner_text).
    """
    wb = load_workbook(file_path, data_only=True)
    ws = wb["MTN Transactions"]

    banner = ws.cell(row=1, column=1).value or ""
    headers = [cell.value for cell in ws[2]]

    data = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        data.append(list(row))

    df = pd.DataFrame(data, columns=headers)
    df["Id"] = df["Id"].astype(str).str.strip()
    if "Date" in df.columns:
        df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    if "Amount" in df.columns:
        df["Amount"] = pd.to_numeric(df["Amount"], errors="coerce")
    wb.close()
    return df, banner


def load_airtel_statement(file_path: Path) -> tuple[pd.DataFrame, str]:
    """Load Airtel consolidated statement Excel.

    Sheet: 'All Transactions'
    Row 0: banner, Row 1: headers, Row 2+: data.
    Returns (dataframe, banner_text).
    """
    wb = load_workbook(file_path, data_only=True)
    ws = wb["All Transactions"]

    banner = ws.cell(row=1, column=1).value or ""
    headers = [cell.value for cell in ws[2]]

    data = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        data.append(list(row))

    df = pd.DataFrame(data, columns=headers)
    if "Transaction ID" in df.columns:
        df["Transaction ID"] = df["Transaction ID"].apply(_normalize_airtel_id)
    if "Transaction Date" in df.columns:
        df["Transaction Date"] = pd.to_datetime(df["Transaction Date"], errors="coerce")
    if "Transaction Amount" in df.columns:
        df["Transaction Amount"] = pd.to_numeric(df["Transaction Amount"], errors="coerce")
    wb.close()
    return df, banner


# ---------------------------------------------------------------------------
# Karibu HMS Ledger CSV
# ---------------------------------------------------------------------------

def parse_karibu_csv(file_path: Path) -> pd.DataFrame:
    """Parse a Karibu HMS ledger CSV.

    Skip first 2 rows (title + blank), row 3 is header.
    Columns: Date, Account, Narration, DR, CR, Balance
    Skip 'Opening Balance' rows.
    """
    df = pd.read_csv(file_path, skiprows=2, dtype=str, quotechar='"')
    df.columns = df.columns.str.strip()

    # Drop Opening Balance rows
    if "Date" in df.columns:
        df = df[df["Date"].str.strip() != "Opening Balance"].copy()
        df["Date"] = pd.to_datetime(df["Date"].str.strip(), format="%Y-%m-%d", errors="coerce")

    # Clean numeric columns
    for col in ["DR", "CR", "Balance"]:
        clean_col = col.strip()
        # Find the actual column (may have trailing spaces)
        matching = [c for c in df.columns if c.strip() == clean_col]
        if matching:
            actual = matching[0]
            df[clean_col] = pd.to_numeric(
                df[actual].str.replace(",", "").str.strip(), errors="coerce"
            ).fillna(0)
            if actual != clean_col:
                df.drop(columns=[actual], inplace=True)

    if "Narration" in df.columns:
        df["Narration"] = df["Narration"].str.strip()
    if "Account" in df.columns:
        df["Account"] = df["Account"].str.strip()

    # Drop fully empty rows
    df.dropna(subset=["Date"], inplace=True)

    return df


def load_all_karibu(karibu_dir: Path) -> pd.DataFrame:
    """Load and combine all Karibu CSV files from a directory.

    Combines chronologically and deduplicates by Date + Narration + DR amount.
    """
    all_dfs = []
    csv_files = sorted(karibu_dir.glob("*.csv"))
    for f in csv_files:
        try:
            df = parse_karibu_csv(f)
            df["_source_file"] = f.name
            all_dfs.append(df)
        except Exception:
            continue

    if not all_dfs:
        return pd.DataFrame()

    combined = pd.concat(all_dfs, ignore_index=True)
    combined.sort_values("Date", inplace=True)

    # Dedup by Date + Narration + DR
    combined["_dedup_key"] = (
        combined["Date"].astype(str) + "|" +
        combined["Narration"].fillna("") + "|" +
        combined["DR"].astype(str)
    )
    combined.drop_duplicates(subset=["_dedup_key"], keep="first", inplace=True)
    combined.drop(columns=["_dedup_key"], inplace=True)
    combined.reset_index(drop=True, inplace=True)

    return combined
