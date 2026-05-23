"""MoMo Agent Transaction Detailed Report XLSX parser.

Used by the Petty Cash UGX account. Direction is driven by Transaction Type
(CASH_IN / DEPOSIT → IN; TRANSFER → OUT). Amount is normalised to a positive
Decimal; source TRANSFER rows are stored as negative in the export
(e.g. `-192100`) and are normalised here to positive with `direction = OUT`,
so `NormalizedRecord.amount` is always positive across every parser.

Unparseable dates pass through tagged with `audit_flag = UNPARSEABLE_DATE`
(Joash, 2026-05-20) rather than being silently dropped.
"""

from __future__ import annotations

from decimal import Decimal, InvalidOperation
from pathlib import Path

import pandas as pd

from parsers._dates import parse_date
from parsers.types import DIRECTION_IN, DIRECTION_OUT, NormalizedRecord


SHEET_NAME = "Sheet1"

_INFLOW_TYPES = {"CASH_IN", "DEPOSIT"}
_OUTFLOW_TYPES = {"TRANSFER"}

# MoMo exports usually deliver `2026-05-15 13:47` (no seconds). Older
# variants may include seconds.
_MOMO_DATE_FORMATS = (
    "%Y-%m-%d %H:%M",
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d",
)


def parse(path: Path) -> list[NormalizedRecord]:
    """Parse a MoMo Agent xlsx and return NormalizedRecord rows.

    Header is row 1 (no skiprows). Transaction IDs are read as str so the
    large integers don't lose precision through a float roundtrip.

    Some older MoMo exports name the sole sheet `Sheet5` instead of
    `Sheet1` (observed on 2024-12-04). We try `Sheet1` first and fall back
    to the only/first sheet in the workbook.
    """
    p = Path(path)
    df = pd.read_excel(p, sheet_name=_resolve_sheet(p), dtype={"Transaction ID": str})
    source = p.name

    records: list[NormalizedRecord] = []
    for idx, row in df.iterrows():
        txn_type = _stripped(row.get("Transaction Type"))
        if not txn_type:
            continue
        direction = _direction_for(txn_type)
        if direction is None:
            continue

        amount = _to_decimal(row.get("Amount"))
        if amount == 0:
            continue

        date_val, flag = parse_date(
            row.get("Date / Time"), _MOMO_DATE_FORMATS,
            source_file=source, row_index=int(idx), field_name="Date / Time",
        )

        from_acct = _account_str(row.get("From Account"))
        to_acct = _account_str(row.get("To Account"))
        counterparty = from_acct if direction == DIRECTION_IN else to_acct

        records.append(NormalizedRecord(
            source_file=source,
            date=date_val,
            txn_id=_stripped(row.get("Transaction ID")),
            amount=abs(amount),
            direction=direction,
            counterparty=counterparty,
            txn_type=txn_type,
            raw={k: _stringify(v) for k, v in row.items()},
            audit_flag=flag,
        ))
    return records


def _resolve_sheet(path: Path) -> str:
    """Return the sheet name to read. Prefers `Sheet1`, else the only
    sheet in the workbook. Used to handle older exports that name the
    sole sheet `Sheet5` (or similar)."""
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True)
    try:
        sheets = wb.sheetnames
    finally:
        wb.close()
    if SHEET_NAME in sheets:
        return SHEET_NAME
    if len(sheets) == 1:
        return sheets[0]
    # Multiple sheets but no `Sheet1` — heuristic: pick the largest one by
    # reading dimensions. Cheap enough for a one-shot.
    wb = load_workbook(path, read_only=True)
    try:
        best = max(sheets, key=lambda s: wb[s].max_row * wb[s].max_column)
        return best
    finally:
        wb.close()


def _direction_for(txn_type: str) -> str | None:
    if txn_type in _INFLOW_TYPES:
        return DIRECTION_IN
    if txn_type in _OUTFLOW_TYPES:
        return DIRECTION_OUT
    return None


def _to_decimal(value) -> Decimal:
    if value is None:
        return Decimal(0)
    try:
        if pd.isna(value):
            return Decimal(0)
    except (TypeError, ValueError):
        pass
    try:
        return Decimal(str(value).replace(",", "").strip() or "0")
    except (InvalidOperation, ValueError):
        return Decimal(0)


def _account_str(value) -> str:
    if value is None or value is False:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    return str(value).strip()


def _stripped(value) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    return str(value).strip()


def _stringify(value):
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    return str(value)
